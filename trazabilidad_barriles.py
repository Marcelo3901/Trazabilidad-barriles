import streamlit as st
import pandas as pd
import requests
from urllib.parse import quote
import base64
import os
from datetime import datetime
import unicodedata
import hashlib
import io
import json
import re
import time
import uuid
import zipfile
import html as html_lib
from copy import copy
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font

# La validación de estados se realiza justo antes de guardar y nuevamente en Apps Script.

# CONFIGURACIÓN DE LA PÁGINA
st.set_page_config(page_title="Trazabilidad Barriles Castiza", layout="centered")


# ---------- CONFIGURACIÓN DE GOOGLE SHEETS PARA INVENTARIO DE LATAS ----------
SPREADSHEET_ID = "1FjQ8XBDwDdrlJZsNkQ6YyaygkHLhpKmfLBv6wd3uluY"
HOJA_INGRESOS_LATAS = "IngresoLatas"
HOJA_MOVIMIENTOS_LATAS = "VLatas"


def normalizar_clave(valor):
    """Normaliza texto para comparar estilos, lotes y estados sin errores de mayúsculas o tildes."""
    if pd.isna(valor):
        return ""
    texto = " ".join(str(valor).strip().split())
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(caracter for caracter in texto if not unicodedata.combining(caracter))
    return texto.casefold()


def limpiar_cantidades(serie):
    """Convierte cantidades de latas a enteros y elimina separadores de miles."""
    texto = serie.fillna("").astype(str).str.strip()
    texto = texto.str.replace(r"[^0-9-]", "", regex=True)
    return pd.to_numeric(texto, errors="coerce").fillna(0).astype(int)


@st.cache_data(ttl=30, show_spinner=False)
def cargar_hoja_csv(nombre_hoja):
    """Carga una pestaña pública del Google Sheet y conserva el resultado por 30 segundos."""
    url = (
        f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/gviz/tq"
        f"?tqx=out:csv&sheet={quote(nombre_hoja)}"
    )
    try:
        df = pd.read_csv(url)
    except pd.errors.EmptyDataError:
        return pd.DataFrame()

    df.columns = df.columns.astype(str).str.strip()
    return df


def calcular_inventario_latas():
    """
    Calcula inventario por estilo y lote:
    IngresoLatas - VLatas.

    En VLatas:
    - Despacho, Baja o Estado vacío: restan inventario.
    - Devolución: vuelve a sumar inventario.
    """
    ingresos = cargar_hoja_csv(HOJA_INGRESOS_LATAS).copy()
    movimientos = cargar_hoja_csv(HOJA_MOVIMIENTOS_LATAS).copy()

    columnas_necesarias = {"Estilo", "Cantidad", "Lote"}
    faltantes_ingresos = columnas_necesarias.difference(ingresos.columns)
    if faltantes_ingresos:
        raise ValueError(
            f"Faltan columnas en {HOJA_INGRESOS_LATAS}: {', '.join(sorted(faltantes_ingresos))}"
        )

    # VLatas puede estar sin registros, pero debe conservar sus encabezados.
    if movimientos.empty and not columnas_necesarias.issubset(movimientos.columns):
        movimientos = pd.DataFrame(columns=["Estilo", "Cantidad", "Lote", "Estado"])

    faltantes_movimientos = columnas_necesarias.difference(movimientos.columns)
    if faltantes_movimientos:
        raise ValueError(
            f"Faltan columnas en {HOJA_MOVIMIENTOS_LATAS}: "
            f"{', '.join(sorted(faltantes_movimientos))}"
        )

    ingresos["Estilo"] = ingresos["Estilo"].fillna("").astype(str).str.strip()
    ingresos["Lote"] = ingresos["Lote"].fillna("").astype(str).str.strip().str.upper()
    ingresos["Cantidad"] = limpiar_cantidades(ingresos["Cantidad"])
    ingresos["_estilo_key"] = ingresos["Estilo"].map(normalizar_clave)
    ingresos["_lote_key"] = ingresos["Lote"].map(normalizar_clave)
    ingresos = ingresos[
        (ingresos["_estilo_key"] != "")
        & (ingresos["_lote_key"] != "")
        & (ingresos["Cantidad"] != 0)
    ]

    ingresos_agrupados = (
        ingresos.groupby(["_estilo_key", "_lote_key"], as_index=False)
        .agg(
            Estilo=("Estilo", "last"),
            Lote=("Lote", "last"),
            Ingresado=("Cantidad", "sum"),
        )
    )

    movimientos["Estilo"] = movimientos["Estilo"].fillna("").astype(str).str.strip()
    movimientos["Lote"] = movimientos["Lote"].fillna("").astype(str).str.strip().str.upper()
    movimientos["Cantidad"] = limpiar_cantidades(movimientos["Cantidad"])
    movimientos["_estilo_key"] = movimientos["Estilo"].map(normalizar_clave)
    movimientos["_lote_key"] = movimientos["Lote"].map(normalizar_clave)

    if "Estado" in movimientos.columns:
        movimientos["_estado_key"] = movimientos["Estado"].map(normalizar_clave)
    else:
        movimientos["_estado_key"] = ""

    # Al restar un movimiento negativo, una devolución vuelve a sumar existencias.
    movimientos["Movimiento_neto"] = movimientos["Cantidad"]
    es_devolucion = movimientos["_estado_key"].eq("devolucion")
    movimientos.loc[es_devolucion, "Movimiento_neto"] *= -1

    movimientos = movimientos[
        (movimientos["_estilo_key"] != "")
        & (movimientos["_lote_key"] != "")
        & (movimientos["Movimiento_neto"] != 0)
    ]

    movimientos_agrupados = (
        movimientos.groupby(["_estilo_key", "_lote_key"], as_index=False)["Movimiento_neto"]
        .sum()
    )

    inventario = ingresos_agrupados.merge(
        movimientos_agrupados,
        on=["_estilo_key", "_lote_key"],
        how="left",
    )
    inventario["Movimiento_neto"] = inventario["Movimiento_neto"].fillna(0).astype(int)
    inventario["Disponible"] = inventario["Ingresado"] - inventario["Movimiento_neto"]
    inventario["Disponible"] = inventario["Disponible"].astype(int)

    # Solo se ofrecen lotes que todavía tienen unidades disponibles.
    inventario = inventario[inventario["Disponible"] > 0].copy()
    inventario.sort_values(["Estilo", "Lote"], inplace=True, key=lambda s: s.astype(str).str.casefold())
    inventario.reset_index(drop=True, inplace=True)
    return inventario


def validar_existencias_latas(latas_solicitadas, inventario):
    """Valida el total solicitado por estilo/lote, incluso si el mismo lote se repite en varias órdenes."""
    if not latas_solicitadas:
        return ["Debes completar al menos una orden de latas."]

    pedido = pd.DataFrame(latas_solicitadas).copy()
    pedido["_estilo_key"] = pedido["estilo"].map(normalizar_clave)
    pedido["_lote_key"] = pedido["lote"].map(normalizar_clave)
    pedido_agrupado = (
        pedido.groupby(["_estilo_key", "_lote_key"], as_index=False)
        .agg(
            Estilo=("estilo", "last"),
            Lote=("lote", "last"),
            Solicitado=("cantidad", "sum"),
        )
    )

    disponibles = inventario[["_estilo_key", "_lote_key", "Disponible"]].copy()
    comparacion = pedido_agrupado.merge(
        disponibles,
        on=["_estilo_key", "_lote_key"],
        how="left",
    )
    comparacion["Disponible"] = comparacion["Disponible"].fillna(0).astype(int)

    errores = []
    for fila in comparacion.itertuples(index=False):
        if int(fila.Solicitado) > int(fila.Disponible):
            errores.append(
                f"{fila.Estilo} / lote {fila.Lote}: solicitaste {int(fila.Solicitado)} "
                f"y solo hay {int(fila.Disponible)} disponibles."
            )
    return errores


# ---------- SEGURIDAD E IDEMPOTENCIA PARA MOVIMIENTOS DE BARRILES ----------
HOJA_BARRILES = "DatosM"
HOJA_CLIENTES = "RClientes"
FORM_BARRILES_URL = "https://docs.google.com/forms/d/e/1FAIpQLSedFQmZuDdVY_cqU9WdiWCTBWCCh1NosPnD891QifQKqaeUfA/formResponse"
FORM_LATAS_URL = "https://docs.google.com/forms/d/e/1FAIpQLSerxxOI1npXAptsa3nvNNBFHYBLV9OMMX-4-Xlhz-VOmitRfQ/formResponse"
URL_DATOS_BARRILES_PUBLICA = (
    f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/gviz/tq"
    f"?tqx=out:csv&sheet={quote(HOJA_BARRILES)}"
)


def leer_secreto(nombre):
    """Lee un secreto sin impedir que la app arranque cuando aún no existe secrets.toml."""
    try:
        valor = st.secrets.get(nombre, "")
    except Exception:
        return ""
    return "" if valor is None else str(valor).strip()


APPS_SCRIPT_MOVIMIENTOS_URL = leer_secreto("APPS_SCRIPT_MOVIMIENTOS_URL")
APPS_SCRIPT_MOVIMIENTOS_TOKEN = leer_secreto("APPS_SCRIPT_MOVIMIENTOS_TOKEN")
BACKEND_SEGURO_ACTIVO = bool(
    APPS_SCRIPT_MOVIMIENTOS_URL and APPS_SCRIPT_MOVIMIENTOS_TOKEN
)


def normalizar_codigo_barril(valor):
    """Convierte códigos numéricos de Sheets (por ejemplo 30275.0) a texto limpio."""
    if pd.isna(valor):
        return ""
    texto = str(valor).strip()
    if re.fullmatch(r"\d+\.0+", texto):
        texto = texto.split(".", 1)[0]
    return texto


def combinar_columnas(df, candidatos):
    """Combina columnas duplicadas de Google Sheets conservando el primer valor no vacío."""
    salida = pd.Series("", index=df.index, dtype="object")
    for nombre in candidatos:
        if nombre not in df.columns:
            continue
        valores = df[nombre].fillna("").astype(str).str.strip()
        vacios = salida.fillna("").astype(str).str.strip().eq("")
        salida.loc[vacios & valores.ne("")] = valores.loc[vacios & valores.ne("")]
    return salida


def convertir_fechas_sheet(serie):
    """Admite fechas DD/MM/AAAA y números seriales de Google Sheets."""
    texto = serie.fillna("").astype(str).str.strip()
    fechas = pd.to_datetime(texto, dayfirst=True, errors="coerce")
    numeros = pd.to_numeric(texto.str.replace(",", ".", regex=False), errors="coerce")
    seriales = fechas.isna() & numeros.between(20000, 80000, inclusive="both")
    if seriales.any():
        fechas.loc[seriales] = pd.to_datetime(
            numeros.loc[seriales], unit="D", origin="1899-12-30", errors="coerce"
        )
    return fechas


def preparar_datos_barriles(df_original):
    """Deja una sola estructura de columnas aunque DatosM tenga encabezados repetidos."""
    if df_original is None or df_original.empty:
        return pd.DataFrame(
            columns=[
                "Marca temporal", "Código", "Lote", "Estilo", "Estado", "Cliente",
                "Responsable", "Observaciones", "Operación ID", "_estado_key", "_orden_fila"
            ]
        )

    df = df_original.copy()
    df.columns = df.columns.astype(str).str.strip()
    salida = pd.DataFrame(index=df.index)
    salida["Marca temporal"] = combinar_columnas(
        df, ["Marca temporal", "Marca temporal.1"]
    )
    salida["Código"] = combinar_columnas(df, ["Código", "Codigo", "Código.1", "Codigo.1"])
    salida["Lote"] = combinar_columnas(
        df, ["Lote", "lote", "Lote.1", "lote.1"]
    )
    salida["Estilo"] = combinar_columnas(df, ["Estilo", "Estilo.1"])
    salida["Estado"] = combinar_columnas(df, ["Estado", "Estado.1"])
    salida["Cliente"] = combinar_columnas(df, ["Cliente", "Cliente.1"])
    salida["Responsable"] = combinar_columnas(df, ["Responsable", "Responsable.1"])
    salida["Observaciones"] = combinar_columnas(
        df, ["Observaciones", "Observaciones.1"]
    )
    salida["Operación ID"] = combinar_columnas(
        df,
        [
            "Operación ID", "Operacion ID", "Operación ID.1", "Operacion ID.1",
            "ID Operación", "ID Operacion"
        ],
    )

    salida["Código"] = salida["Código"].map(normalizar_codigo_barril)
    salida["Marca temporal"] = convertir_fechas_sheet(salida["Marca temporal"])
    salida["Lote"] = salida["Lote"].fillna("").astype(str).str.strip()
    salida["Estilo"] = salida["Estilo"].fillna("").astype(str).str.strip()
    salida["Estado"] = salida["Estado"].fillna("").astype(str).str.strip()
    salida["Cliente"] = salida["Cliente"].fillna("").astype(str).str.strip()
    salida["Responsable"] = salida["Responsable"].fillna("").astype(str).str.strip()
    salida["Observaciones"] = salida["Observaciones"].fillna("").astype(str).str.strip()
    salida["Operación ID"] = salida["Operación ID"].fillna("").astype(str).str.strip()
    salida["_estado_key"] = salida["Estado"].map(normalizar_clave)
    salida["_orden_fila"] = range(len(salida))
    return salida


def cargar_csv_publico_fresco(nombre_hoja, timeout=30):
    """Consulta la hoja sin caché para validar el estado real antes de guardar."""
    nonce = int(time.time() * 1000)
    url = (
        f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/gviz/tq"
        f"?tqx=out:csv&sheet={quote(nombre_hoja)}&_={nonce}"
    )
    respuesta = requests.get(
        url,
        timeout=timeout,
        headers={"Cache-Control": "no-cache", "Pragma": "no-cache"},
    )
    respuesta.raise_for_status()
    try:
        df = pd.read_csv(io.StringIO(respuesta.text))
    except pd.errors.EmptyDataError:
        return pd.DataFrame()
    df.columns = df.columns.astype(str).str.strip()
    return df


@st.cache_data(ttl=10, show_spinner=False)
def cargar_datos_barriles_ui():
    """Carga rápida para construir la interfaz; al guardar siempre se consulta de nuevo."""
    return preparar_datos_barriles(cargar_csv_publico_fresco(HOJA_BARRILES))


def cargar_datos_barriles_frescos():
    return preparar_datos_barriles(cargar_csv_publico_fresco(HOJA_BARRILES))


def obtener_ultimos_movimientos(df):
    """Obtiene la última fila registrada de cada código, usando el orden real de la hoja."""
    if df is None or df.empty:
        return pd.DataFrame(columns=df.columns if isinstance(df, pd.DataFrame) else [])
    validos = df[df["Código"].astype(str).str.strip().ne("")].copy()
    if validos.empty:
        return validos
    validos.sort_values("_orden_fila", inplace=True)
    return validos.drop_duplicates(subset="Código", keep="last").copy()


def obtener_ultimo_movimiento(df, codigo):
    codigo = normalizar_codigo_barril(codigo)
    if not codigo or df is None or df.empty:
        return None
    historial = df[df["Código"] == codigo].sort_values("_orden_fila")
    if historial.empty:
        return None
    return historial.iloc[-1]


def barriles_actualmente_en_cuarto_frio(df):
    ultimos = obtener_ultimos_movimientos(df)
    if ultimos.empty:
        return ultimos
    return ultimos[ultimos["_estado_key"].eq("en cuarto frio")].copy()


def validar_formato_codigo_barril(codigo):
    codigo = normalizar_codigo_barril(codigo)
    if not re.fullmatch(r"(?:20|30|58)\d{3}", codigo):
        return False, "El código debe tener 5 dígitos y empezar por 20, 30 o 58."
    return True, ""


def operacion_ya_registrada(df, operacion_id):
    if not operacion_id or df is None or df.empty:
        return False
    op = str(operacion_id).strip()
    if "Operación ID" in df.columns and df["Operación ID"].eq(op).any():
        return True
    patron = f"[OP:{op}]"
    return df["Observaciones"].fillna("").astype(str).str.contains(
        re.escape(patron), regex=True
    ).any()


def validar_movimiento_barril_local(df, codigo, nuevo_estado, estilo="", lote=""):
    """
    Reglas principales:
    - nunca permite repetir consecutivamente el mismo estado;
    - Despacho solo se autoriza si el último estado real es En cuarto frío;
    - En cuarto frío exige estilo y lote.
    """
    valido, mensaje = validar_formato_codigo_barril(codigo)
    if not valido:
        return False, mensaje, None

    ultimo = obtener_ultimo_movimiento(df, codigo)
    estado_nuevo_key = normalizar_clave(nuevo_estado)

    if ultimo is not None and ultimo.get("_estado_key", "") == estado_nuevo_key:
        cliente_anterior = str(ultimo.get("Cliente", "")).strip()
        extra = f" para {cliente_anterior}" if cliente_anterior and estado_nuevo_key == "despacho" else ""
        return (
            False,
            f"El último estado del barril ya es '{nuevo_estado}'{extra}. "
            "El registro repetido fue bloqueado.",
            ultimo,
        )

    if estado_nuevo_key == "despacho":
        if ultimo is None:
            return (
                False,
                "Este barril no tiene historial. Solo se puede despachar un barril cuyo último estado sea 'En cuarto frío'.",
                None,
            )
        if ultimo.get("_estado_key", "") != "en cuarto frio":
            estado_actual = str(ultimo.get("Estado", "Sin estado")).strip() or "Sin estado"
            cliente_actual = str(ultimo.get("Cliente", "")).strip()
            detalle_cliente = f" (cliente: {cliente_actual})" if cliente_actual else ""
            return (
                False,
                f"No se puede despachar. El estado actual es '{estado_actual}'{detalle_cliente}; "
                "debe estar actualmente en 'En cuarto frío'.",
                ultimo,
            )

    if estado_nuevo_key == "en cuarto frio":
        if not str(estilo).strip():
            return False, "Debes seleccionar el estilo para ingresar el barril al cuarto frío.", ultimo
        if not str(lote).strip():
            return False, "Debes ingresar el lote para ingresar el barril al cuarto frío.", ultimo

    return True, "", ultimo


def construir_huella_operacion(datos):
    serializado = json.dumps(datos, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha256(serializado.encode("utf-8")).hexdigest()


def obtener_operacion_id_para_huella(huella):
    op_anterior = st.session_state.get("movimiento_operacion_id", "")
    huella_anterior = st.session_state.get("movimiento_huella", "")
    if op_anterior and huella_anterior == huella:
        return op_anterior
    nuevo = uuid.uuid4().hex
    st.session_state.movimiento_operacion_id = nuevo
    st.session_state.movimiento_huella = huella
    return nuevo


def limpiar_operacion_pendiente():
    st.session_state.pop("movimiento_operacion_id", None)
    st.session_state.pop("movimiento_huella", None)


def movimiento_local_reciente(codigo, estado, segundos=600):
    """Capa adicional para cubrir el retraso con que Google Forms refleja una respuesta."""
    if not codigo:
        return False
    confirmados = st.session_state.get("movimientos_confirmados_localmente", {})
    llave = f"{normalizar_codigo_barril(codigo)}|{normalizar_clave(estado)}"
    registro = confirmados.get(llave)
    if not registro:
        return False
    return (time.time() - float(registro.get("momento", 0))) <= segundos


def marcar_movimiento_local_confirmado(codigo, estado, operacion_id):
    if not codigo:
        return
    confirmados = dict(st.session_state.get("movimientos_confirmados_localmente", {}))
    ahora = time.time()
    confirmados = {
        k: v for k, v in confirmados.items()
        if ahora - float(v.get("momento", 0)) <= 1800
    }
    llave = f"{normalizar_codigo_barril(codigo)}|{normalizar_clave(estado)}"
    confirmados[llave] = {"momento": ahora, "operacion_id": operacion_id}
    st.session_state.movimientos_confirmados_localmente = confirmados


def enviar_backend_apps_script(payload):
    """Envía una operación atómica. El Apps Script vuelve a validar bajo LockService."""
    cuerpo = dict(payload)
    cuerpo["token"] = APPS_SCRIPT_MOVIMIENTOS_TOKEN
    cuerpo["accion"] = "registrar_movimiento"
    try:
        respuesta = requests.post(
            APPS_SCRIPT_MOVIMIENTOS_URL,
            json=cuerpo,
            timeout=45,
        )
    except requests.RequestException as exc:
        return False, f"No se recibió respuesta del servidor seguro: {exc}", True, {}

    try:
        datos = respuesta.json()
    except ValueError:
        return (
            False,
            f"El servidor seguro devolvió una respuesta no válida (HTTP {respuesta.status_code}).",
            respuesta.status_code >= 500,
            {},
        )

    if datos.get("ok"):
        mensaje = datos.get("message") or "Operación registrada correctamente."
        return True, mensaje, False, datos

    mensaje = datos.get("message") or "La operación fue rechazada por el servidor seguro."
    reintentable = bool(datos.get("retryable", False))
    return False, mensaje, reintentable, datos


def esperar_confirmacion_operacion_formulario(operacion_id, intentos=3, pausa=2):
    for _ in range(intentos):
        time.sleep(pausa)
        try:
            datos = cargar_datos_barriles_frescos()
        except Exception:
            continue
        if operacion_ya_registrada(datos, operacion_id):
            return True
    return False


def enviar_por_formularios_compatibilidad(payload):
    """
    Modo compatible con la app antigua. Evita dobles clics por sesión e incluye un ID
    en Observaciones, pero no sustituye el bloqueo transaccional de Apps Script.
    """
    operacion_id = payload["operacion_id"]
    codigo = payload.get("codigo", "")
    registrar_barril = bool(codigo)

    if registrar_barril:
        try:
            datos = cargar_datos_barriles_frescos()
            if operacion_ya_registrada(datos, operacion_id):
                return True, "La operación ya estaba registrada; no se volvió a enviar.", False, {"duplicate": True}
        except Exception:
            pass

        observaciones = str(payload.get("observaciones", "")).strip()
        etiqueta = f"[OP:{operacion_id}]"
        observaciones_con_id = f"{observaciones}\n{etiqueta}".strip()
        datos_formulario = {
            "entry.311770370": codigo,
            "entry.1283669263": payload.get("estilo", ""),
            "entry.1545499818": payload.get("estado", ""),
            "entry.91059345": payload.get("cliente", ""),
            "entry.1661747572": payload.get("responsable", ""),
            "entry.1465957833": observaciones_con_id,
            "entry.1234567890": payload.get("lote", "") if payload.get("estado") in ["Despacho", "En cuarto frío"] else "",
            "entry.1122334455": payload.get("incluye_latas", "No"),
            "entry.1437332932": payload.get("lote", ""),
        }
        try:
            respuesta = requests.post(FORM_BARRILES_URL, data=datos_formulario, timeout=25)
        except requests.RequestException as exc:
            if esperar_confirmacion_operacion_formulario(operacion_id):
                return True, "El registro sí llegó a Google Sheets y no se duplicó.", False, {"verified_after_error": True}
            return False, f"No se pudo confirmar si Google recibió el barril: {exc}", True, {}

        if respuesta.status_code not in (200, 302):
            return False, f"Google Forms rechazó el barril (HTTP {respuesta.status_code}).", False, {}

    errores_latas = []
    for indice, item in enumerate(payload.get("latas", []), start=1):
        datos_latas = {
            "entry.457965266": str(item["cantidad"]),
            "entry.689047838": item["estilo"],
            "entry.2096096606": item["lote"],
            "entry.1478892985": payload.get("cliente", ""),
            "entry.1774006398": payload.get("responsable", ""),
            "entry.1179145668": "Despacho",
        }
        try:
            respuesta_latas = requests.post(FORM_LATAS_URL, data=datos_latas, timeout=25)
        except requests.RequestException as exc:
            errores_latas.append(f"Orden {indice}: {exc}")
            continue
        if respuesta_latas.status_code not in (200, 302):
            errores_latas.append(f"Orden {indice}: HTTP {respuesta_latas.status_code}")

    if errores_latas:
        return (
            False,
            "El barril pudo quedar registrado, pero fallaron órdenes de latas: " + "; ".join(errores_latas),
            True,
            {"partial": registrar_barril},
        )

    return True, "Operación registrada correctamente.", False, {}



def crear_id_suboperacion(operacion_general_id, tipo, identificador):
    """Genera un ID estable para que cada barril del pedido pueda reintentarse sin duplicarse."""
    semilla = f"{operacion_general_id}|{tipo}|{identificador}"
    return uuid.uuid5(uuid.NAMESPACE_URL, semilla).hex


def registrar_lista_despacho(barriles, datos_comunes, operacion_general_id):
    """
    Registra todos los barriles del pedido con un solo clic.

    Cada barril se envía como un movimiento independiente a Registro_Barriles/DatosM.
    Las latas, cuando existen, se incluyen únicamente en el primer movimiento para evitar
    que se repitan por cada barril. Si el despacho no tiene barriles, se envían solas.
    """
    latas_pedido = list(datos_comunes.get("latas", []))
    resultados = []

    movimientos = []
    for indice, barril in enumerate(barriles):
        incluye_latas_en_este = indice == 0 and bool(latas_pedido)
        payload = {
            "codigo": barril["codigo"],
            "estado": "Despacho",
            "cliente": datos_comunes.get("cliente", ""),
            "responsable": datos_comunes.get("responsable", ""),
            "observaciones": datos_comunes.get("observaciones", ""),
            "lote": barril.get("lote", ""),
            "estilo": barril.get("estilo", ""),
            "incluye_latas": "Sí" if incluye_latas_en_este else "No",
            "latas": latas_pedido if incluye_latas_en_este else [],
            "operacion_id": crear_id_suboperacion(
                operacion_general_id, "barril", barril["codigo"]
            ),
        }
        movimientos.append((f"Barril {barril['codigo']}", payload))

    if not barriles and latas_pedido:
        payload_latas = {
            "codigo": "",
            "estado": "Despacho",
            "cliente": datos_comunes.get("cliente", ""),
            "responsable": datos_comunes.get("responsable", ""),
            "observaciones": datos_comunes.get("observaciones", ""),
            "lote": "",
            "estilo": "",
            "incluye_latas": "Sí",
            "latas": latas_pedido,
            "operacion_id": crear_id_suboperacion(
                operacion_general_id, "latas", "despacho_solo_latas"
            ),
        }
        movimientos.append(("Despacho de latas", payload_latas))

    for etiqueta, payload in movimientos:
        if BACKEND_SEGURO_ACTIVO:
            exito, mensaje, reintentable, extra = enviar_backend_apps_script(payload)
        else:
            exito, mensaje, reintentable, extra = enviar_por_formularios_compatibilidad(payload)
        resultados.append({
            "etiqueta": etiqueta,
            "exito": bool(exito),
            "mensaje": mensaje,
            "reintentable": bool(reintentable),
            "extra": extra or {},
            "payload": payload,
        })

    exitosos = [r for r in resultados if r["exito"]]
    fallidos = [r for r in resultados if not r["exito"]]
    detalles = [
        f"{r['etiqueta']}: {'registrado' if r['exito'] else r['mensaje']}"
        for r in resultados
    ]

    if not fallidos:
        cantidad_barriles = len(barriles)
        if cantidad_barriles and latas_pedido:
            mensaje = (
                f"Se registraron {cantidad_barriles} barril(es) y "
                f"{len(latas_pedido)} orden(es) de latas."
            )
        elif cantidad_barriles:
            mensaje = f"Se registraron {cantidad_barriles} barril(es) del pedido."
        else:
            mensaje = f"Se registraron {len(latas_pedido)} orden(es) de latas."
        return True, mensaje, False, {"resultados": resultados, "detalles": detalles}

    mensaje = (
        f"Se registraron {len(exitosos)} de {len(resultados)} movimientos. "
        "Los movimientos fallidos pueden reintentarse sin volver a seleccionar el pedido."
    )
    return (
        False,
        mensaje,
        any(r["reintentable"] for r in fallidos),
        {"resultados": resultados, "detalles": detalles, "partial": bool(exitosos)},
    )

# ---------- ORDEN GENERAL DIARIA Y FORMATOS F-TRZ ----------
ZONA_HORARIA_COLOMBIA = ZoneInfo("America/Bogota")
NOMBRES_PLANTILLA_FORMATOS = (
    "FORMATOS_DE_DESPACHOS.xlsx",
    "FORMATOS DE DESPACHOS.xlsx",
    "FORMATOS DE DESPACHOS(1).xlsx",
)

# Copia interna de la plantilla. Permite generar los formatos aunque el archivo XLSX
# no haya sido subido al repositorio o Streamlit ejecute la app desde otra carpeta.
PLANTILLA_FORMATOS_EMBEBIDA_BASE64 = """
UEsDBBQABgAIAAAAIQBrK/rSlgEAAK8GAAATAAgCW0NvbnRlbnRfVHlwZXNdLnhtbCCiBAIooAACAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADEVctOwzAQvCPxD5GvKHHbA0KoKQceR0ACPsCNt4lpYlveBdq/Z+OmFUJ9ELUSl7zsnZmd9W7GN4umTj4h
oHE2F8NsIBKwhdPGlrl4e31Ir0SCpKxWtbOQiyWguJmcn41flx4w4WiLuaiI/LWUWFTQKMycB8srMxcaRfwaSulVMVclyNFgcCkLZwkspdRiiMn4Dmbqo6bk
fsGfV0qmxorkdrWvpcqF8r42hSIWKj+t/kWSutnMFKBd8dEwdIY+gNJYAVBTZz4YZgwvQMSJoZBbOd99+YvTNK3mdw/ljpAANfbT2RmRcWTMBSvj8YLd2sHQ
ruw2oot74goGoyF5VoEeVcN2yUUtv1yYT52bZ/tB+roZXc0aZexa9x7+uBllvA1PLKTNLwL31DH6Jx3E7QEyXo+3IsIcSBxpWQOeuvwR9BBzpQLoF+LGK08u
4Cf2AR06qK9Wguwejve9A+rJe/yR+wMvT7/n4DzygA3Qv+rr0dRGp56BIJCBzXDa1uQbRp7ORx8zaMe/Br2FW8bfzeQbAAD//wMAUEsDBBQABgAIAAAAIQC1
VTAj9AAAAEwCAAALAAgCX3JlbHMvLnJlbHMgogQCKKAAAgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAArJJNT8MwDIbvSPyHyPfV3ZAQQkt3QUi7IVR+gEncD7WNoyQb3b8nHBBUGoMDR3+9fvzK2908jerIIfbiNKyLEhQ7I7Z3rYaX+nF1ByomcpZGcazh
xBF21fXV9plHSnkodr2PKqu4qKFLyd8jRtPxRLEQzy5XGgkTpRyGFj2ZgVrGTVneYviuAdVCU+2thrC3N6Dqk8+bf9eWpukNP4g5TOzSmRXIc2Jn2a58yGwh
9fkaVVNoOWmwYp5yOiJ5X2RswPNEm78T/XwtTpzIUiI0Evgyz0fHJaD1f1q0NPHLnXnENwnDq8jwyYKLH6jeAQAA//8DAFBLAwQUAAYACAAAACEAw7PHlygD
AAAvBwAADwAAAHhsL3dvcmtib29rLnhtbKxV227bOBB9L9B/EPQuk9RdQpzCtmw0QFsEqbcF9qWgKToiLIkqScUOgvx7h3Lk3PYh29aQSZEjHp6ZOUOefTg0
tXPDlRaynbpkgl2Ht0yWor2euv+sV17qOtrQtqS1bPnUveXa/XD+/t3ZXqrdRsqdAwCtnrqVMV2OkGYVb6ieyI63YNlK1VADQ3WNdKc4LXXFuWlq5GMco4aK
1j0i5OotGHK7FYwXkvUNb80RRPGaGqCvK9HpEa1hb4FrqNr1ncdk0wHERtTC3A6grtOw/OK6lYpuanD7QCLnoOCJ4U8wNP64E5hebdUIpqSWWzMBaHQk/cp/
ghEhz0JweB2DtyGFSPEbYXN4YqXi32QVn7DiRzCC/xiNgLQGreQQvN9Ei07cfPf8bCtq/u0oXYd23Rfa2EzVrlNTbZalMLycugkM5Z4/m1B9N+9FDVY/yDBx
0flJzpfKKfmW9rVZg5BHeKiMOM78yH4JwpjVhquWGr6QrQEdPvj1p5obsBeVBIU7V/xnLxSHwgJ9ga/QUpbTjb6kpnJ6VR8jqKHkykkpmZ7U4oZPWm7QJkhL
whI/Y1EU8zJEBdc7Izu0oDXra6nRE+nS13XyP8RLmY0IgpAcaR/fX4YH2Kt8FOilUQ68XxSfIElf6Q2kDIQBHgwVfQE5SX/cpXG8mKer1MuKJPXCJIy92TyI
vXS+XM1m8zjDwewevFBxziTtTfUgA4s5dUMr3Jemz/QwWgjOe1E+7n+HH36e7V80o+3eemoPvG+C7/WjYOzQOXwXbSn3U9cjPnhz+3y4H4zfRWkqUBxOAvjk
OPeRi+sKGBNCYpikzEAS13QDM9YF3/KcunfEj4poESbeKs4CL0zTwkuL5cILSYpnPvGXSYQHfugJweGgBaJD77RDcay89dW/4CAcXsO0DTfspHK7jbooh1L4
rwXBkwWw+LTAH/I/bsVAYFBAthuQM4L9zH7BD+aTNkMP2hXWpRDPEpyFHl4GEbiU+V4aBr63CAt/GSXLYjmPbIbt5ZL/jSN2KKF8vLUsy4oqs1aU7eCuu+Lb
OdWgxSECCPiClEfWaFx1/gsAAP//AwBQSwMEFAAGAAgAAAAhAEqppmH6AAAARwMAABoACAF4bC9fcmVscy93b3JrYm9vay54bWwucmVscyCiBAEooAABAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAALySzWrEMAyE74W+g9G9cZL+
UMo6eymFvbbbBzCxEodNbGOpP3n7mpTuNrCkl9CjJDTzMcxm+zn04h0jdd4pKLIcBLram861Cl73T1f3IIi1M7r3DhWMSLCtLi82z9hrTk9ku0AiqThSYJnD
g5RUWxw0ZT6gS5fGx0FzGmMrg64PukVZ5vmdjL81oJppip1REHfmGsR+DMn5b23fNF2Nj75+G9DxGQvJiQuToI4tsoJp/F4WWQIFeZ6hXJPhw8cDWUQ+cRxX
JKdLuQRT/DPMYjK3a8KQ1RHNC8dUPjqlM1svJXOzKgyPfer6sSs0zT/2clb/6gsAAP//AwBQSwMEFAAGAAgAAAAhAP8FsoV5CgAA1zYAABgAAAB4bC93b3Jr
c2hlZXRzL3NoZWV0MS54bWyck9uO2jAQhu8r9R0s3+d8IEGEFSxFXakXVdXDtXEcYhHHqW0WVlXffcchASQkhFYCxiT6v5l/Zjx7OooGvTKluWwLHLg+Rqyl
suTttsC/fq6dDCNtSFuSRraswG9M46f550+zg1Q7XTNmEBBaXeDamG7qeZrWTBDtyo618KaSShADf9XW051ipOxFovFC3089QXiLT4SpeoQhq4pTtpJ0L1hr
ThDFGmKgfl3zTo80QR/BCaJ2+86hUnSA2PCGm7ceipGg05dtKxXZNOD7GMSEoqOCTwjfaEzTP7/JJDhVUsvKuED2TjXf2s+93CP0TLr1/xAmiD3FXrkd4AUV
fqykIDmzwgss+iAsPcNsu9R0z8sC/1tmeRAslqGTZnHgxCs/dvJkuXSyOE0nfhotVnn2H89nJYcJW1dIsarAi2D6LUqxN5/1C/Sbs4O+OiO7jxspd/bFC+Tx
AaFZw6jdDEQgvLJn1jQF/hLYnf7bU+0ZkN6ZeX0e+et+h78rtCGaPcvmDy9NDZcF7krJKrJvzA95+Mr4tjbwNLFAKhtQwy8S3N4rWCdy7OPhJM7dZOJHQZhg
tGHarLmVYkT32kgxJhhIJwYMpGdAHBjhxI3DZJL1lDtK6H6vhDgoo+iS/o4wHoQQB6H1/GC94KzPCnEUx26YJUGSWtd30qaDEuLo9Kpdd4STsc1Q5aB8MGVw
nhAcxnKv23vftNfP+x0AAP//AAAA//+cmu1u20YURF/F0APEppTEceEYqCiR3I+XCIIg6Y82Reymr9+ldqi93MNaEf8YxsHcJTncuzsi+fj87cuXl8Onl09P
jz++/3vz4+Om2dw8//3pr+f032/NdnPz7SX917zZvtvcfP7n+eX7n8OXP76e4Obp8fNY8ftY8nGTFLcie5H3m5tU/ZyEP5/uHm9/Pj3efpakleT+XHQAOYJ0
ID3IAOJAvMiH89FDJunv+Yyb+RlH1Ywlt8mss2PJpGsdG0uSrQ/FsUy2hbTSGINAjiAdSA8ygDgQn8nurhgkjTFoWxmka4BBu+sNGkvmBmViDZLGGARyBOlA
epABxIH4TKxB0hiDdpVBugYY9PZ6g8aS1HPl/uwz2TXnO9aKbEuLgRxBOpAeZABxIF5kV2ZQJrbF3lYG6apgUFp80GK7tLr834o06pM7ye9zN7+bH2q/IHlf
LVELkvu55LAg+TCXHBckD3NJtyBpqvWyz5pdcqgsUNUKNUhTFmK3VFW1rUdVWDqfei5Lg1uV7gpu1cOb+1e2j7FiPpVBWpADyBGkA+kzsfOiqebgsCCp5o67
PIq/PErA2UUReHq/tD68e9XUsSSZ+rbsMCJlNWgzKZJDBmX6HGvQ1aDPoCkb6QDiQDxIAIk6X5jx4XozxpL5bgLSZmLMyMCYUYOuBj1GHUAciAcJIDGT072b
ZY+H680YS+btBtJmYszIwJhRg64GPUYdQByIBwkgUQQzo7m73o1TTbKjbJv7CZlOETKGiBhHQDqQXsS2C5Ej8kSBKE5nTmcWU/3rK0ijRGxCKlErZJ3JddaZ
mnSqKpqeQw9EjsgTBaIoxAYaf+IwvV9wRkHXhK/TMLOuaoWsM7nOOlOTTlXWGRxt4NEckScKRHFCnDOLsf2CM0q4tpuEbDdlZJ3JxDpTk66pSS8y6yaF8LIf
Oao8USCKQgtzZjGvX3BG4dd2E1DbZGSdycQ6U5NOVXbOYOhBIvOb0xF5okAUhRacWQrqzQVnlCRtNwG1TUbWmUysMzXpVGWdwdCDROa3lCPyRIEoTojdtJSL
LzmjPGi7Sch2U0bWmUysMzXpmpr0IrNuyiKDHFWeKBBFoYU5syLdNoqLtpuAWqmsM0i40hSvOpCeRxuIHJEnCkRRaMGZFVE33as63hG1QtYZxF1prDMIvBx6
IHJEnigQxQmxm1bk3rTynZyx3SRkuwnRV3W2mxB+oelFZt2Uy2bdBORZGIjidDFwJq1iV+eZU031OHIcZoZaqcycETHOgHQgPY82EDkiTxSIohC7KUX8653R
M1+zN52GmSc9IesMMrA0pptAeg49EDkiTxSI4oQ4Z1Zk4K1SqemmCZluErLOIANLY52pNb00tpuIHJEnCkRxOnM6syIDj08d64f7QK1U1hlkYGmsM8jAPNpA
5Ig8USCKQgvdtCIDp6dR9d5E1ApZZ5CBpbHO1JqeQw9EjsgTBaI4Ic6ZFRl4fLs2PpSx3SRkuwkZWHV2BUYGhqYXmXVTLrN7E1WeKBDF6WLozIoMvFUEta/K
gFqp7JxBBpbGzhlkYB5tIHJEnigQRaGFblqRgbd6QGr3JqBWKusMMrA01hk85+XRBiJH5IkCUZwQ58yKDLxVBrbdJGS7KSPrDDKwRrLOIANLM+smPcA1TyGo
8kSBKE4XQ2dWZOD0+hR7E1ArlXWmTrxHaawztabn0QYiR+SJAlEUYjell7JXJ71TzfyxOFErZJwRMSswSAfSc+iByBF5okAUJ4Q5k94OX++MMrDpptMwySzT
TULWGWRgacycAelFbDcROSJPFIjidOZ0ZkUG3umTCLM3EbVC1hlkYGmsM8jAHHogckSeKBBFoYVuWpGBd3roa/YmolbIOoMMLI11BhmYQw9EjsgTBaI4Ic6Z
xQx89+aVd93j+/vxx0FKHuU1fvXdwX4SlS9bWqH34zmcPro6iBRNJ5L2njJ09b1Cf9ZM4wwg7hfG8agKIFFkYUYtJuTXfcv5NG0x5dqqDy32O2XYMunaCZW1
6yBUfDuCdCIzJ6tvNvqzpjiZj282/V8Yx2OcABJFFpxcTNSvO5lDr3VyW31rsh8/rRlnqXVSyDqZkXWyJp1Gsk5uq69W+rOmOJnHsU5OpNz/ehyPcQJIFDFO
3pZPIP8DAAD//wAAAP//bNXtbtowGAXgW4l8AYMkDSmvCJINrb+0X72CDAxEowkK6Sr16nfKRDdp5198HmT5OMRevabxmDbpfL5mu+GtnxpVV2q9+oqzMR0a
pXPRD2r2X25yCSzXlegF+30lhuWbSjYs31ayZflTJU8sf67kmeW2EleR9fhKAstjJZHNo2vRS9arFsNyW4tluavFsdzX4lkeagksj7VEmudziXlBVqrzUnTO
OhuIoWIhloqDOCoe4qkESKASIZGKRh9N+xiIoWIhloqDOCoe4qkESKAS8wVW/Uj3eim6yOl3sxRDxeZLsVQcxFHxEE8lQAKVCIlUNPpo2sdADBULsVQcxFHx
EE8lQAKVWBRYNT1xChw5Rc32GmKoWIil4iCOiod4KgESqERIpKLRR9M+BmKoWIil4iCOiod4KgES+I6WD2JKttexeJRYztk/vsRFUZbsLUAMFQuxVBzEUfEQ
TyVAApUIiVQ0+mjax0AMFQuxVBzEUfEQTyVAwh+Z/b2R16tLe0zf2/HY9dfsnA64neffapWN3fF0f56Gyy2tVPZjmKbh9T46pXafxs9RqbLDMEz3Aa7xz3lf
0vR2yS7tJY0v3UdqFGYYxi71Uzt1Q9+oc9vvrzu4yk6AjwFy3l66RpWLucp+pXHqdv8mo3T7Ro1+fzvy9mP73vXH7Cu9XUez92H8eT2lNK1/AwAA//8DAFBL
AwQUAAYACAAAACEAhHUxNRoJAACJKAAAGAAAAHhsL3dvcmtzaGVldHMvc2hlZXQyLnhtbJyT246bMBCG7yv1HZDvOWNIopBVNi1qpF5UPV4bY4IVjKntHFbV
vnsHJ2QjpdpGKwFjjP/vnxmb+cNRtM6eKc1ll6PQC5DDOior3m1y9ON74U6Qow3pKtLKjuXoiWn0sHj/bn6QaqsbxowDhE7nqDGmn/m+pg0TRHuyZx18qaUS
xMCr2vi6V4xUViRaPwqC1BeEd+hEmKl7GLKuOWUfJN0J1pkTRLGWGMhfN7zXI03Qe3CCqO2ud6kUPSBK3nLzZKHIEXS23nRSkbKFuo9hQqhzVHBFcMejjZ2/
cRKcKqllbTwg+6ecb8uf+lOf0Avptv67MGHiK7bnwwa+oKK3pRTiCyt6gcVvhKUX2NAuNdvxKkd/0hgX2WO4cjHOAjfJpthdLj9m7hI/TlYFXhVRMH1Gi3nF
YYeHqhzF6hwtw9k6xshfzO0B+snZQV+NHUPKb6xl1DAwCZEznM9Syu2wcA1TASC1XTAgCTV8z1asbXP0eTjiv60JDMHAvzhcj0e3wp7oL8opiWYr2f7ilWnA
Ev6citVk15qv8vCJ8U1jYNZmTGULang6gsNfBq0V5Gjj4SSOsBcmQRph5NCdNlKM0CGdiw66aHUQz7o49pIIZ5PwP8rkrIR4VoaBN8E4SSfZ657w1XqmF+Xr
67Pzeoij0+Q+J9gF6wRxVGZX1ZVMm4IPLf1Xj3zb4r8AAAD//wAAAP//lJnrbts6EIRfxdADxJHk3AonQE2bVh8jCIK2P3pBndPz+ofSjsRZDY9g/UrwYZfW
iEPuktpfvr2/fxxfP15f9n9+/bv581zV1eby+/XnJf336anafPtI/7Q3d9Xm7Z/Lx68f3fv3rwOrXvZvffznPuG52jXVFuRgpLmvNin5kgL/vjTNfvv3Zb99
Q0xAzMOUdRRyEhKFnEEep3E6I3VDv15PP75NKiepKWSd1D4hvY6nLNVIk0lADAkTchIShZyNtLdZGGJYWNOWlbVrlfUJXpkRVoYYUibkJCQKORthZYhxynZl
Zbu1yvqE56rJ7/FgpK2nNxtAsoWPQk5CopAzSJvnzIgz4/8ISytsbsbmbmnd9Qlp3aV3N62yO7/IDghJD5EX4iwmICatminmwQ9zREj6M4U8+pATQtySv/cx
0WLa/HbOyHrg55v9eFd4vmb2618sZpg+t8DT48gCr2+ahc2sz/BuMcJukZgjYrKyk5GaJ6d5mr0PxPDstLc+5oyYvL908utfQHrTOfXptYr6Bel9eJK+y5sb
SF4WwUgOORq4m3JOBtr7iUSQvG2cJabDuMMidRoe12now/02JiQYIQ0GSIPkRCO0GZ6NtFR3QFRDqqJr5qEP9xYUEoyQBgOkQXKiETLyGYTMBTIY2c1DfbtO
xBCfVORN9jAishMQ6QAhISDsqBGRpTSqGwfXCalLPc5Sh4OWgur+MIRzWgBiMZbHYmSkiDT2FhCba0QFMaUuZkkMWhYqiH158pYLQCzGgliMpEWksclGRC4b
UcFmpcZlSQy6FLYZENvMEIsxwmLQldDG1e/e6bW0bDOJ6hC1K8xMqVdZEoNmgW0mKNSGWIwRFiNpEWnOZmhXuHcGKojpa+z8XLAkBjWZbSYo1KjudAowwmIk
LSLN2cyiWrYZUMFmpcZgSQxKLNsMiG1miGfGCItBq8A2A2KbSVRXY/DCzHCdT0VoOJ0tibFiy6eYWlAAYjFS6zUtAjmbodyzzYAKYrjgXyMGlZltJijUUvNB
eGYkLSLI2QxVnm0GVLAZV/5rxKBos82A2GZS/NNU9hsVi0EhZ5sBsc0kqsNQhd0snaDyBnCFmCF+dljuh/BFE1FkMxASoyNFILYZEBfNEanNUl+yTgyuGchm
wxCuTwtALEY6AE2LQGyzEZHNRqQ2a7gDuGZmULfJZsMQSQzZDIjFSAeAGO7NRkQ206huHLwwM9wBXCMGdwd8JyMoNNIBgLDNJC0iyNkMHQDtZoiyCw1/wcQd
wDVirPzyZUU6j817MyCeGekANC0COZuh3LPNgAo24w7gGjGo22wzILaZdAD98dzvZiDOZqjtbDOgvOd1SCztZtwBXCNGDuKH/n5zvptJB4AgtpmkRQQ5m6ED
YJsBFdbMyg6gwcGedzNBAVFsM+kAdKQI5GyGcs82AyrYbGUH0KBus82A2GbSASCPZwaFnIomgvgIMCK2GQYvzMzKDiBZQG6YBQVE8cxIB6AjRSBnM3QAbDMg
FZM2uFVFc4j3NxmKAhCJAaGZ0bQIxDYbEdlsRGqzlLhODDoAstkwhC+aQCxGOgDE8G42ItrNNKobBy/MTKkDWPxo0+JjApVNRQGI5UgPoGkRiI0GxN3ZiApy
Sj3Ashwc+mlHS/fM88IJxHLkHkDTIpCzGroAthpQwWrFLuDmYeGM1loVTpvkdOPe5o9Yw3e3A2LSWXj88BYccp1IWyzey89g9dQ9w+wr3mEYNnX0/AyWZsie
YZu/Mv4HAAD//wAAAP//bNVbjpswGAXgrVheQBOghMmvEMlOfOlDn2YFNHES1Awgh+lIs/qeSZVppZ43fD6DDubizUvK57RL1+tNHcbXYW71qtLbzWescjq1
2hRivurFf7ktJLDc1GJWbH4tluW7WnYs39eyZ7mrxdfk+qGWwObHWiLLTSNmzXo2YlnuGnEs9414lodGAstjI5HmRSWxYHdmiqWYoqTPYCmWisM5joqHeCoB
EqhESKRi0NrQ1hZiqTiIo+IhnkqABCqxWEssC7I6plih2xNdtxW6MXE4x1HxEE8lQAKVCIlUDFob2tpCLBUHcVQ8xFMJkEAllvgwyoatW1miG/3mIZaKgzgq
HuKpBEigEiGRikFrQ1tbiKXiII6Kh3gqARKoxKqQWFV03Z7EVEv2vpVPYqk4iKPiIZ5KgAQqERKpGLQ2tLWFWCoO4qh4iKcSIOGPLP7uL9vN1J3T9y6f++Gm
rumEvWb5pdEq9+fL43gep3taa/VjnOfx5TG6pO6Y8seo0uo0jvNjgE3p47rPaX6d1NRNKT/376nVa63G3Kdh7uZ+HFp97Ybj7QDX6gJ4HyHX/dS3ulottfqV
8twf/k2y9MdW52/H+2/lmLu3fjirz/T+K168jfnn7ZLSvP0NAAD//wMAUEsDBBQABgAIAAAAIQCDTWzIVgcAAMggAAATAAAAeGwvdGhlbWUvdGhlbWUxLnht
bOxZW48bNRR+R+I/WPOe5jaTy6opyrVLu9tW3bSIR2/iZNz1jCPb2W2EKqHyxAsSEiBekHjjASGQQALxwo+p1IrLj+DYM8nYG4de2CJAu5FWGec7x8fnHH8+
c3z1rYcJQ6dESMrTTlC9UgkQSSd8StN5J7g3HpVaAZIKp1PMeEo6wYrI4K1rb75xFe+pmCQEgXwq93AniJVa7JXLcgLDWF7hC5LCbzMuEqzgUczLU4HPQG/C
yrVKpVFOME0DlOIE1I5BBk0Juj2b0QkJrq3VDxnMkSqpByZMHGnlJJexsNOTqkbIlewzgU4x6wQw05SfjclDFSCGpYIfOkHF/AXla1fLeC8XYmqHrCU3Mn+5
XC4wPamZOcX8eDNpGEZho7vRbwBMbeOGzWFj2NjoMwA8mcBKM1tcnc1aP8yxFij76tE9aA7qVQdv6a9v2dyN9MfBG1CmP9zCj0Z98KKDN6AMH23ho167N3D1
G1CGb2zhm5XuIGw6+g0oZjQ92UJXoka9v17tBjLjbN8Lb0fhqFnLlRcoyIZNdukpZjxVu3ItwQ+4GAFAAxlWNEVqtSAzPIE87mNGjwVFB3QeQ+ItcMolDFdq
lVGlDv/1JzTfTETxHsGWtLYLLJFbQ9oeJCeCLlQnuAFaAwvy9Kefnjz+4cnjH5988MGTx9/mcxtVjtw+Tue23O9fffzHF++j377/8vdPPs2mPo+XNv7ZNx8+
+/mXv1IPKy5c8fSz75798N3Tzz/69etPPNq7Ah/b8DFNiES3yBm6yxNYoMd+cixeTmIcY+pI4Bh0e1QPVewAb60w8+F6xHXhfQEs4wNeXz5wbD2KxVJRz8w3
48QBHnLOelx4HXBTz2V5eLxM5/7JxdLG3cX41Dd3H6dOgIfLBdAr9ansx8Qx8w7DqcJzkhKF9G/8hBDP6t6l1PHrIZ0ILvlMoXcp6mHqdcmYHjuJVAjt0wTi
svIZCKF2fHN4H/U48616QE5dJGwLzDzGjwlz3HgdLxVOfCrHOGG2ww+win1GHq3ExMYNpYJIzwnjaDglUvpkbgtYrxX0m8Aw/rAfslXiIoWiJz6dB5hzGzng
J/0YJwuvzTSNbezb8gRSFKM7XPngh9zdIfoZ4oDTneG+T4kT7ucTwT0gV9ukIkH0L0vhieV1wt39uGIzTHws0xWJw65dQb3Z0VvOndQ+IIThMzwlBN1722NB
jy8cnxdG34iBVfaJL7FuYDdX9XNKJEGmrtmmyAMqnZQ9InO+w57D1TniWeE0wWKX5lsQdSd14ZTzUultNjmxgbcoFICQL16n3Jagw0ru4S6td2LsnF36Wfrz
dSWc+L3IHoN9+eBl9yXIkJeWAWJ/Yd+MMXMmKBJmjKHA8NEtiDjhL0T0uWrEll65mbtpizBAYeTUOwlNn1v8nCt7on+m7PEXMBdQ8PgV/51SZxel7J8rcHbh
/oNlzQAv0zsETpJtzrqsai6rmuB/X9Xs2suXtcxlLXNZy/jevl5LLVOUL1DZFF0e0/NJdrZ8ZpSxI7Vi5ECaro+EN5rpCAZNO8r0JDctwEUMX/MGk4ObC2xk
kODqHarioxgvoDVUNc3OucxVzyVacAkdIzNsmqnknG7Td1omh3yadTqrVd3VzFwosSrGK9FmHLpUKkM3mkX3bqPe9EPnpsu6NkDLvowR1mSuEXWPEc31IETh
r4wwK7sQK9oeK1pa/TpU6yhuXAGmbaICr9wIXtQ7QRRmHWRoxkF5PtVxyprJ6+jq4FxopHc5k9kZACX2OgOKSLe1rTuXp1eXpdoLRNoxwko31wgrDWN4Ec6z
0265X2Ss20VIHfO0K9a7oTCj2XodsdYkco4bWGozBUvRWSdo1CO4V5ngRSeYQccYviYLyB2p37owm8PFy0SJbMO/CrMshFQDLOPM4YZ0MjZIqCICMZp0Ar38
TTaw1HCIsa1aA0L41xrXBlr5txkHQXeDTGYzMlF22K0R7ensERg+4wrvr0b81cFaki8h3Efx9Awds6W4iyHFomZVO3BKJVwcVDNvTinchG2IrMi/cwdTTrv2
VZTJoWwcs0WM8xPFJvMMbkh0Y4552vjAesrXDA7dduHxXB+wf/vUff5RrT1nkWZxZjqsok9NP5m+vkPesqo4RB2rMuo279Sy4Lr2musgUb2nxHNO3Rc4ECzT
iskc07TF2zSsOTsfdU27wILA8kRjh982Z4TXE6968oPc+azVB8S6rjSJby7N7VttfvwAyGMA94dLpqQJJdxZCwxFX3YDmdEGbJGHKq8R4RtaCtoJ3qtE3bBf
i/qlSisalsJ6WCm1om691I2ienUYVSuDXu0RHCwqTqpRdmE/gisMtsqv7c341tV9sr6luTLhSZmbK/myMdxc3VdrztV9dg2PxvpmPkAUSOe9Rm3Urrd7jVK7
3h2VwkGvVWr3G73SoNFvDkaDftRqjx4F6NSAw269HzaGrVKj2u+XwkZFm99ql5phrdYNm93WMOw+yssYWHlGH7kvwL3Grmt/AgAA//8DAFBLAwQUAAYACAAA
ACEAlwwfs04FAACxKgAADQAAAHhsL3N0eWxlcy54bWzUWt1v4jgQfz/p/oco72k+SlhAhFVpG2mlvdVJ7Un7sC8mccA6J0GO6cGe7n+/sRMgbDEkaUIpD5AY
e+bn+fKM7fHndUy1F8wykiaebt9YuoaTIA1JMvf0v559Y6BrGUdJiGiaYE/f4Ez/PPn9t3HGNxQ/LTDmGpBIMk9fcL4cmWYWLHCMspt0iRP4J0pZjDi8srmZ
LRlGYSYGxdR0LKtvxogkek5hFAdViMSI/b1aGkEaLxEnM0IJ30hauhYHoy/zJGVoRgHq2u6hQFvbfeZoa7ZlIltf8YlJwNIsjfgN0DXTKCIBfg13aA5NFOwp
AeVmlGzXtJyDua9ZQ0o9k+EXItSnT8bJKvZjnmlBuko4qHPXpOX/fAmhsd/TtVwr92kIcgp/GPEPYwMf3ZyMzYLGZBylSYnULdASLZNx9lN7QRQI2aJ/kNKU
aRx0DqRkS4JinPe4R5TMGBHdIhQTusmbHdEgzaToFxNQmmSec8i/Z6JXZV53jCB6msbgHN7jNA4xHJ2dWca962+1we+UfM/Pefg2CF0PLxlIQxGqKBwYT+eK
ODsPafJ1VCnFkYHfEEp3/uwIH4SGyRhCH8cs8eFFK56fN0vwwASidO4Fst+Z3nOGNrbjlgaYkuFkPEtZCKvCLpK4wDpvm4wpjjj4JiPzhfjl6RK+ZynnEDon
45CgeZogKmLJdkR5JCwnsHJ4Ol9A5N/GDzafebrvW/Ij4QgmBY+KIyQeCafiAADeKe5cNp2AORR/pyw+qvQvgLsii9w1rsIzunXU+sFga8e1lXVlEWcH591N
4hcdtIxHEfY74tLcKq5EDfUD87VI8l2QF/kCpD0BpvRJ5Anfo4NqZh2VKhmoVUXeLYoa8QgJT/GYpxv5i0hDytRy2iWyvQFkIfXpautox0A1WhRax1C5UEIW
ozW0XNKND9MQ1VP+NpW51/79jpJ5EuNtRQdFVf4q6nZOAlGLBfAvzkuodaSeDjBWiOkkoPYA9DuWyCJl5CcIsyST11LS/mFo+YzXUuTCPk6J7NNH0eG7GRXs
EB01KtgfqGjmoIO6LgQOf9yFTtBSqbIJLZVrK2nlrq30pEqG29RWq2uiBkpRoOmtuFKnhmKDpqoF2xOWY6vWl+rUL6F/Jcy2YkMrVmo7VxUwapt87VVXnZxU
SSJUa3bl6HoJy7tVqBTW+zc7n4r2sCntYgc8zxwri/fbKp5h5svjjH1u1jBz6zbct5Retp9Mgc0eTxWuC7GI6tvK4gPbx1VI+1zOoloMoL2VGqkVT1OFIGjv
LLy1kbeocEMm3BluKAc6oy0SnO6It1GvqCQukp43I1dVz3ZlO2wjF6iV8yvl0Upa/g4R7lw8E8XM0TWuw4DWGFPl9Ky+3TTGVDmtax8TsD6qusq22oGYlJXn
1SZ5Kil2WLCcNTaVFDt0gPq7s8ryvfJyffHwrlqRKicBV4O4kz2yYlv+jVvQSru4qPUeXfbl0QocppRObA7Oa3YnL5q4N+Xp30T5TEuha7YilJPkyFkN0AzX
+9MfeYeHi5t98lxoxwU8O8QRWlH+vPvT0/fPf+CQrGJYfItef5KXlEsSnr5//irusth9cfMEzgW+ZnD1BH61FSOe/u/j9NPw4dF3jIE1HRi9W+waQ3f6YLi9
++nDgz+0HOv+P5iTuAY5gvuFb7hdKK9Dwm643RtlFO4gsmKyBfinfZunl15y+PLeDMAuYx86fevOtS3Dv7Vso9dHA2PQv3UN37Wdh35v+uj6bgm72/AWomXa
dn6fUYB3R5zEmJJkq6uthsqtoCR4PTEJc6sJc3/XdPI/AAAA//8DAFBLAwQUAAYACAAAACEAAbXEKwsCAADNBQAAFAAAAHhsL3NoYXJlZFN0cmluZ3MueG1s
jFTBatwwEL0X+g9C94292VKK8TpQF0OhbJZN0kNvWmtqD8gaV5KXJn/TD+gpn7A/VtmmTWI5xeCLpTdPM2/eTHr1s1HsBMYi6S1fX8ScgS5Joq62/O62WH3g
zDqhpVCkYcvvwfKr7O2b1FrHfKy2W1471yZRZMsaGmEvqAXtb76TaYTzv6aKbGtASFsDuEZFl3H8PmoEas5K6rTb8ncbzjqNPzrIx4PNhmepxSx1WTHwEJPA
9iBRkmX3TKJ1Bo9diedH3V+1hmRXOmIOTINaSEojl6VRzzHy7A1VRjSC7Ve3h2+rdTwF5OdHiRUlrBgAcXw5RXztdfIPJixeB/TnX5V/l637bKa3vVKJbUXp
FfRSWDAn4FkBZS3YlGhHzdGAZ1EsVwjawQQxT+bLGwWYEs7Dc6EdSiGXob+Qg4W8o4hD+h+FMaiW5qOGdPqy//VyUeEKmxbhQQxPnqA+/y47RctevT72rRCl
Nz/YoBN50Jrpyby2B7AtaSuOamyjM0L7E7NYQ9InP4TzSR1AKHwQSeDehOVd06qpW7KbrvU1og1DdnnCdjQf9opjn2aRjIRh8rxDDVS9/K82LgtGa/Of0Qpm
az6XcE78XviEBspxKQQDOiwPxnIyHuOC/fDCC9PgAv0y60v0n4/Go99DU8zN58Av14sc/Lelyzz71M9n+Miv4+wPAAAA//8DAFBLAwQUAAYACAAAACEAtX47
TloCAABtBAAAGAAAAHhsL2RyYXdpbmdzL2RyYXdpbmcxLnhtbJxUXW+bMBR9n7T/YPmd8BFIExRSJSFMlaot6jbt2TUmeMM2st00bdX/vmsT2q2TpmlPOfjm
Xp9z7oHl5Ul06Mi04UoWOJ5EGDFJVc3locBfv1TBHCNjiaxJpyQr8AMz+HL1/t3yVOv83pQawQBpcngscGttn4ehoS0TxExUzyRUG6UFsfCoD2GtyT2MFl2Y
RNEsNL1mpDYtY7YcKvg8j/zHNEG4xCvPDLhuWdetJW2VHo4arcSAqOpW0TJ0Chz0DQA+Nc0qvZgl2UvJnfiqVvdjh4Pj2ZsOKPkOP/n1OnayiJ4KnGVJmoG/
9KHA6SKbRhEOh1E9pwOQxz2n+zNh+vG414jXBU4wkkSA+1yQA4sn3/sD6CQ5TL429ozQneYFfqqqZJPtqjSoAAVptEmDzS5dBFUyne+Si2qbTGfPrjue5RTc
t7D4q3p0PZ794bvgVCujGjuhSoSqaThl4x5hi3Eaet890afdej3PpotFkG2TXZBuy0Uwj9dpsIsvZps4yublJnp2skPPfvz1KoaFONHnlQB0biAIScP0DeuA
7JHdMMMfwQvvnu/5zbXbjvcV72CtJHf4rOyf0jmIKxW9E0zaIaLaX6ukaXlvMNI5E7cMdqKv6hh2Ce+GBTK95tI6XSQ3VjNLWwcb4HHDqB30vhQ86VeeLoem
d6pJfmo0hJTkwARBZHwcMXKBccEcLvh7nsDYcUivjf3AlAADjQXGwMTHhhzB74HT+Bd3p1TON3f+ysjDl3jSjoMvJbEENdeK/jDfuG0/u7f313W8eff8CPel
WP0EAAD//wMAUEsDBAoAAAAAAAAAIQDMpUPjLB8AACwfAAATAAAAeGwvbWVkaWEvaW1hZ2UxLmpwZ//Y/+AAEEpGSUYAAQIAAAEAAQAA/9sAQwAIBgYHBgUI
BwcHCQkICgwUDQwLCwwZEhMPFB0aHx4dGhwcICQuJyAiLCMcHCg3KSwwMTQ0NB8nOT04MjwuMzQy/9sAQwEJCQkMCwwYDQ0YMiEcITIyMjIyMjIyMjIyMjIy
MjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIy/8AAEQgAyQDJAwEiAAIRAQMRAf/EAB8AAAEFAQEBAQEBAAAAAAAAAAABAgMEBQYHCAkKC//E
ALUQAAIBAwMCBAMFBQQEAAABfQECAwAEEQUSITFBBhNRYQcicRQygZGhCCNCscEVUtHwJDNicoIJChYXGBkaJSYnKCkqNDU2Nzg5OkNERUZHSElKU1RVVldY
WVpjZGVmZ2hpanN0dXZ3eHl6g4SFhoeIiYqSk5SVlpeYmZqio6Slpqeoqaqys7S1tre4ubrCw8TFxsfIycrS09TV1tfY2drh4uPk5ebn6Onq8fLz9PX29/j5
+v/EAB8BAAMBAQEBAQEBAQEAAAAAAAABAgMEBQYHCAkKC//EALURAAIBAgQEAwQHBQQEAAECdwABAgMRBAUhMQYSQVEHYXETIjKBCBRCkaGxwQkjM1LwFWJy
0QoWJDThJfEXGBkaJicoKSo1Njc4OTpDREVGR0hJSlNUVVZXWFlaY2RlZmdoaWpzdHV2d3h5eoKDhIWGh4iJipKTlJWWl5iZmqKjpKWmp6ipqrKztLW2t7i5
usLDxMXGx8jJytLT1NXW19jZ2uLj5OXm5+jp6vLz9PX29/j5+v/aAAwDAQACEQMRAD8A96SNCikqpOB2p/lp/cX8qE+4v0FOoAb5af3F/Kjy0/uL+VOooAb5
af3F/Kjy0/uL+VOooAb5af3F/Kjy0/uL+VOooAb5af3F/Kjy0/uL+VOooAb5af3F/Kjy0/uL+VOooAb5af3F/Kjy0/uL+Vctd+ONGs/Gdr4Yku4vttxC8n3+
FcEbYzxgMw3nBI+6P7wrq6AG+Wn9xfyo8tP7i/lTqKAG+Wn9xfyo8tP7i/lTqKAG+Wn9xfyo8tP7i/lTqKAG+Wn9xfyo8tP7i/lTqKAG+Wn9xfyo8tP7i/lT
qKAG+Wn9xfyqjmtCs+gC8n3F+gp1NT7i/QU6gAooooAKKKKACiiigAooooAKzodQs7y8vbGKeOW4tSqXMAOWj3KGXcPQg8djz6GtGvln4ieJNV8KfHPV9W0q
4MNwnkKwIykiGCLKOO6nH5gEYIBABv6r8FdUu/Flzq1rpVnFoi6gpXSkudsk1uGG8qR8q7hkgbgQDjjAFe5z3GnaHa2UDtb2cLSR2ltEAEBY8JGij+Q6AE9A
a4TS/i3aXfwxvfF8+mypNZSfZ5bVHGHmO3btY9FO9SSRkc8NgZ8b0zxvrHjX4r+G7zVZwVj1CFYLeMYjhUyAkKPU8ZJyTgc8DAB9Z0UUUAFFFFABRRRQAUUU
UAFFFFABWfWhWfQBeT7i/QU6mp9xfoKdQAUUUUAFFFFABRRRQAUUVSv72303T7q+upPLgtommmfaTtRQSTgcnAB6UAea/HXU7zRvCmk6jp9xJbXcGrRtHLG2
Cp8mYfkRkEdCCQa5DTNZ8M/GjTk0vxKItO8Vwp5dpeRDaJx1GAeDz1jJ7kqRk7el+PhN/wDDOxvLTFxbC/hnM0Pzp5bRyAPuHG0llAPQ7h618xAkHIoA+hLD
4beJLL4S674XazjfUbjVUeEpIPLkjHlZcMcYHyt1APHTOKqsnhf4J6OCfs+seN3T5e6WxYdf9hQD7O+f4QfltaD8SPEQ+BepazJcpLqWn3S2UN1Km5yp8vDN
nhnG88nrgZyck+CXFxNd3EtxcSvNNKxeSSRizOxOSSTyST3oA+iPgRr2p+JtT8U6jq1091cv9lG5uir++wqjooGTwPWvba8D/Ztt5lt/EVw0Uggd7eNJCp2s
yiQsAehIDLkdtw9a9q0vVbHWLNL7Trlbq1kd0SaPlWKMUbB7jcp5HB6jI5oA0qKKKACiiigAooooAKKKKACs+tCs+gC8n3F+gp1NT7i/QU6gAooooAKKKKAC
iiigArn/AB3/AMk98S/9gq6/9FNXQVz/AI7/AOSe+Jf+wVdf+imoA+YvAnxNv/CZGmX6/wBpeHJiyXFjIA+1W+8U3cdzlT8rZPQncOm8X/Cqy1vTz4o+Hckd
7p8iFpbCJiXRgMnywec46xn5gemcgDxeuq8F+ONX8Daut9p8geCQBbm1kJ8udff0YdmHI9wSCAdXog/4xt8Sg/8AQYj/APaNSeBvhWl5p6eKPGE8emeHYx52
yRtj3Cdueqoxxg/eYfd+8Gr1uw8ZeC9R8AX3iyXTkisVuRLfWzW4ZmuxtwMfddiSmG6cgkqQceB+PviPq3ju7jF0BbabAxaCyjYlQem5j/E2OM9ucAZOQDa8
f/E/+1ID4c8Kx/2Z4agUxBIk8trn1JA+6nX5epyS3JwvtvwT/wCSR6F/28f+lElfIFfX/wAE/wDkkehf9vH/AKUSUAegUUUUAFFFFABRRRQAUUUUAFZ9aFZ9
AF5PuL9BTqan3F+gp1ABRRRQAUUUUAFFFFABXH3utaX4qXxJ4NS9FtqawSWsiSAE7ZI+JEGRvADcjIII5wCCewr5B+Kt1cWXxg1m6tZ5ILiKeN45YmKshEaY
II5BoAw/FfgvWvBmoCy1a2CeZkwzod0UwHUq3txkHBGRkciuar6L8HfEnQfiLp0XhbxtaQG9lISOR1Cx3L9FIIx5cpyemASTjGQtcB8Q/hNqfgmSTUbUtfaG
ZMLOB+8gB6CUD8tw4J/ukgUAW9E/5No8R/8AYZj/APaFeUV6von/ACbR4j/7DMf/ALQrC8A/DbV/Hl07wEWmmwkCW8lUlc/3EH8TY5x0A6kZGQDC8N+G9U8V
avFpej2zTXD/ADMeiRqOruewH8yAMkgV9QaBNpfwy8LeHfC+q6lHNqU8y28MUK/NI80xOQCeEUvyxxkL0yQtcz4h8XeGfg1pY8PeGrGK41gxgybjkqccPOww
SxzkIMcH+EFc+NeHdXv9Z+Kmh3+o3kl1dXGr2rSSueSfNXj0AHQAcAcDigD7NooooAKKKKACiiigAooooAKz60Kz6ALyfcX6CnU1PuL9BTqACiiigAooooAK
KKKACvjn4xf8lY17/rrH/wCikr7Gr45+MX/JWNe/66x/+ikoA4avc/hr8Z3j26D4yn+0Wcq+XFfzDcUzxtmz95T/AHuo75ByvhlFAH2Snw+8MXXhy90q2iYa
Rqd2moPBby4jJ+Q4QjojbAcA9ztIGMedfET4xWWkaePDvgeSFCg8uS8t0AigQcbIccE/7Q4A6ZJyvPaJqV9F+zX4hCXlwAmoi3QCQ/LE3lbkHop3NkDg7j6m
vHKALM08t1PJPNI8s0jF3d2LMzE5JJPUk1r+BP8AkoXhr/sK2v8A6NWqGnaTqGr3DW+mWFxezKhdo7aFpGCggE4UE45HPvWh4E/5KH4Z/wCwra/+jVoA+3qK
KKACiiigAooooAKKKKACs+tCs+gC8n3F+gp1NT7i/QU6gAooooAKKKKACiiigArg/iB8NNJ8dWm+RRaarGP3V7GgJI/uuP4l/UdjyQe8ooA+H/E3hfVvCOry
abq1qYnUnZKuTHMv95G7j9R0IByKwK+4/EvhrS/FWjy6VqtsJoH+ZW6PE46Oh7MP1BIOQSD8v/EL4Xap4JupJ4kkvdHYgx3qj/V5ONsgH3TnjPRsjGDlQAam
if8AJtHiP/sMx/8AtCvKK9X0T/k2jxH/ANhmP/2hXN+Bvh7rXjm/C2MZh0+OQJcX0g+SLjJAH8TY/hHqMkA5oA9y/Z/0m2s/AL6jGgNzf3LmVz12odqr9B8x
+rGrvw5+E2neC9mo3TJf6yUwZiv7uDPURA9+248kdNoJB67wt4ZsPCegwaPp3nG3iyd0rbmZicknt19ABW/QAUUUUAFFFFABRRRQAUUUUAFZ9aFZ9AF5PuL9
BTqan3F+gp1ABRRRQAUUUUAFFeEfHTx/daXd2fh/RdQntbpcXF5PazFHQEEJHuVsjIJYgjpsI6mvGP8AhO/GH/Q165/4MJf/AIqgD7eorzL4K+KtR8T+CJG1
W4NxdWVybYTMPneMIpUuf4m5Iz1OBnJyT2PibxLpnhTRZtW1SfyYI/lVRy8rnoiDuxx+ABJwASADcqrcW8N1bS29xEk0MqlJI5FDK6kYIIPBBHavl7xb8cPE
uuzSw6XL/Y9geFSAgzMOOWkxkHIP3dvBwc9a7TwV4j1hfgN4n1fUdUv5rpZJ0t7qa5dpEzHGq7WJyMOTjB65oA72P4XeHYfDl54dhS4j0q8vFvJYBLkgrs+V
WPIU7BnqeTgjjHU6Xpdno+nQ2FhbxQWsC7Y4o02qozn8ycknuTmvjL/hO/GH/Q165/4MJf8A4qvRvhh8YNVsNat9I8R6jJe6ZdMI1ubqTMls5PDFzyUJODuP
AwQRgggH0xRXgvx+1LWtD1bRbzS9d1KyjuoJIngtrp4kzGwO7CkZJ8zH/ARXjf8AwnfjD/oa9c/8GEv/AMVQB9vUV8+/CjxDrHjHw54k8L3WuXf9qtD9psLx
7qbzkPCn94DkIrCPgdd7dcmsX4WfEHW9N+IEekeJNT1GeC8c2Tx3s0khgnzhDhskHcNhHA+bJ+7QB9OUUUUAFFFFABRRRQAVn1oVn0AXk+4v0FOpqfcX6CnU
AFFFFABWRr+s2vh7QL3V79isFpEZHwRlvRRnA3E4A9yK16+d/wBoHxgZ7y38J2sn7u323N5ju5HyIfop3dwdy9xQBwXhvS7/AOKXxK23zyZvZmub2WPJ8qIc
kKTnAxhFzkDKipPin4U0XwZ4mh0rR7u4nxbLLcC4kR2jck4U7QMfLtOCM4IPeuT03WdS0aZ59L1G7sZnXYz2szRMy5zglSOMgcV0fgLw0fH3jiOx1DUjH5m6
5uZZX3TTgEFlQnq5znJ6AM3OMEA90+B+mNoXwza+v3S3ju55L3MvyeXEFVQWJxgYQtnphga8O+JHjm48d+JXuk3pp1tmKygY/dTu5HZmxk+gwMnbmvZPjn4j
Xw54ItPDWmKtu18PJ2RfKIraMDKjBGM/KuMEFd4rjPgN4Kj1fW5vEl7EHttNcJbK6ghp8Z3fVBtPTqykHigDy/XPD2oeHZoLbVIRb3M1utx5BPzxqxOA4/hb
AztPIBGcHivaXjeD9kzDKVYgNgjGQb7IP5EV5p8WdQi1P4pa9PCGCJOIMH+9Eixt+GUNe0eMU8v9mOBNuwjTNPyuMYO+HP60AeP/AAo8G6f458UXWm6lNdQw
xWTXCtbsqtuDouPmVhjDnt6Vl+PvDVv4S8ZajotrO89vbshjeXG/DIrYbHGRux2z1wKxtO1bUdHuGuNM1C7sp2TY0ltM0TFcg4JUg4yBx7Cor6+utSu5Lu9u
Zrm5kxvmnkLu2BgZY8ngAfhQB9AfFOK58Q/AnQNbkKT3EC2lzcznAb549j4+ruuQPT2rzv4TeCvD/jnUdQ07WLm/guYYUngW1ZV3Jkq5YsrDgtHjp1PWvb9Q
0q01P4DtYQTC+gTREaCS2U/vnijDIQCM8ug+UjPY818s6dq2paRcNcaXfXVlOyFGktpmiYrkHGVIOMgcewoA+pvCvwe0Pwhr0OsaZqWsGdFZDHLNGY5FYYIY
KgJHQ9eoB7V498dPDC6F43/tSCMJa6uhmG0AATLgSDAOecqxJ6lz6Gn/AAl+IOtJ8QbC01fWr67s74Nalby5klVXblCoJOGLhVz6Ma9m+Lnhc+J/AN5DCm+9
sv8AS7cAEligO5QACSShYAd220AXPhp4s/4TPwPZ6jI7Nexf6Nd8f8tkAy3QD5gVfA4G7Hauzr5U+B/iz/hH/Gi6ZcSAWWrhbc/7MwP7o9CeSSuOB8+T0r6r
oAKKKKACiiigArPrQrPoAvJ9xfoKdTU+4v0FOoAKKKKAMHxVr8Hhfw1f63cktHaxFljyR5jnhFyAcZYgZxxnJ4FfI+h6bqHj7xxb2sspe51O6aS5nAUYBJeR
8cDgBiBx0wK9D+PnjEanrMPhi2lDW2nnzrkjB3TleB/wFSeh6uQR8teU6Jr+qeHbx7vR72WzuHjMTSRHkqSDj8wPyoA9N+Knww8OeBvDltfadfX0l5NcCJYb
qWNgybWLEBVU8ELzyOcdxVD4CaW1/wDEyK7DlV0+1lnPy5DFh5e32/1hP/Aa4bUdY1zxTqEP2+8vdTu3YRQI7NI2SQAqL7nHAHJr6M8EeGf+FU/DrVtW1FYj
qjwtdXCeZhRsUmOHPQnJIyByXIGQAaAPGPi9r7a98SNT+ZjBYt9hhDKAVEZIYcdR5m8gnnBH0qfwr8YNc8H6DDo2mabo5gRmcySQyGSRmOcsVcAnoOnQAdq4
uwsrjXdctbGKRTdX1ykKvMxwXdgAWPJ6nk813fiL4KeIvC+gXes6hqGkfZbVQziOaQsckKAAYwMkkDr3oA4DUr+fWNXvNRuMG4vJ3nk2DA3uxY4Hpk19SfHG
VLT4UXsCYVZZYIlUegcNj/x2vm3wPGsvj/w7GyB1bU7YMpGQR5q5r6A/aKdh8PbIA4DapGD7jypT/SgDxz4VeENM8ceKZ9L1Oa5hiSzedGtmVWLh0GPmVhjD
N29KvfFr4dWHgGfSm068up4b5Jcpc7SyMhXJ3KACDvHGOMdTnjhtG1zUvD94b3Sb2W0uShQyRHB2nqP0FT614m1rxG0Umsalc3rQ5EfnPkJnrgdBnAz64HpQ
B7z+zrrM174a1PR5CzR6fOkkTM2QFlDfIB2AZGb3Lmvn/W9NbR9e1HS3cSNZXUlszqMByjFc47ZxXvP7OekXVvpesatNCVt7t4obdzkbvL37yPUZZRn1BHav
N/jRpZ034o6oywCKG8Ed1Fj+PcgDN+Lh/wAaAPVNA+B/hG5s9M13TtU8QQmWOK8t38+EOmQHU/6vgjivZ6+H7fxj4ntLdLe18R6tBBEoWOKK9lVUUdAAGwBW
54W+JPiPSfE+m32o69qt1YxTr9phnupJlaI8P8jNgnaSR6HBoAh+JPho+EPHl9YwqY7V3FzZlRtAickgDkn5TlM9Tsz3r6c+HnitfGHgyx1RmDXYXybsDHEy
4DHA4GeGx2DCuM+PvhhNT8KQa/Co+1aW+JD3aFyAegySG2kcgAFzXztpviHW9HSSPTNY1Cxjc7nW1uXiDH1IUjNAH3VRXzJ8IviTqtv4wi07XtWvb6z1Pbbo
bqd5fKmz8hGckZJ2nGB8wJPy19N0AFFFFABWfWhWfQBeT7i/QU6mp9xfoKdQAVQvL2KyspLqRZmSIjiC3eZ+TjhEBY9ew6c9Kv1natqcejaLfapOjvFZ28lx
IqY3FUUsQM45wKAPCrr4V+Fb27mu7rVvF8txPI0ksj6DdEuxOST+46kmmxfCHwUsgM194wdO4TQrlT+Ztz/Kuuj+L93Fpl7can4We0ni0qLV7WP7eki3Fu8i
oCWC5Q5YEAgng5xxm/efETxHpehrqWpeCfs0k91b21nD/akb/aDKG53Kp242r1HO7tigCHwno/gbwawl03R9bkvQMG9udEvZJu/Q+TheGIO0DIxnNaXi+70b
xd4budEuB4gt4LkoZHj0G8LYVgwAzD6qKpW/xSkMlvaX2hm01P8Atm30m6s2uw/2fzgSkocJtcfK3Ax069KNR+J1xbJdyWHh/wC2GLX10G3RrwRNPPtJY8oQ
o3BQMnnOSRjFAHK+Gvh74Q8OeIrLWVn8TXb2j+YkM2gXewtg7TxD2OCPcCu68WXui+LfDV3olyviCCG52bpI9CvCw2urjGYcdVFVdT8e69p76Vp6+ETLr15B
NcTaf/aSAW8cbYDeZt2tuGTgYI6VSg+Kmo61cabb+GvCx1WW90sag6/2gtv5P71onT50w211xnvnOMUAc74Z+H/hLw34is9ZjuPE909o5dIptAu9pbaQCcQ9
ic/UV1XjmLQvHuiQ6XeHxFbRxXK3AeHQbwkkKy45h6fOfyqDXvileaJrGrW0vhWSfTtFa1/tC8jv0DRCdVIxGVG45Yjg4OOSAavp4pWPxN45vHjvHPh6yj2w
td/uZF8t5TtjCgKxK4LNuPTBA4IB5v8A8Kf8Hf8AQS8Xf+E/df8Axirdj8KPAkE6SXcni26VGDeV/Y10iuAeVbEGcH2IPvXrvgrxKfGHhKx177J9k+1eZ+58
zzNu2Rk+9gZ+7np3rPsfG/2/4mat4NXT9jWFotz9r87O/IjO3Zt4/wBb1yenSgCxYeIdB0uxisrDTdYt7aJdscUeg3oVR/35riPHfhjw1471mDVLyfxNbSxW
4t9sOhXZUqGZgeYeuWP6VoaV8Vr2+vbT7V4Za10y7v3063vRfrJvuBnahTaCNxGMngZpLX4neIJLrXIrvwV9mXRLaSe/b+1o38oiF5EXAT5t20DK5xnJ6YoA
4f8A4U/4O/6CXi7/AMJ+6/8AjFH/AAp/wd/0EvF3/hP3X/xivQE+JGtJ4Xi1i98H/Z2v5bWLSYRqcb/bWnyRlgv7vC4PzDnOOKuWHxEMvhrxHqepaWum32hs
6XFk92HDMEBQCQLjLE7RgHn1yKAJrXUdGh8LRaBc2+uXVotkLKQtol6plj2bOcRdSvUjv6V5b/wp/wAHf9BLxd/4T91/8Yr0p/ijYR+DPDvieW0aO01e9Szl
3ygfZSd4dicfMqtGfTI546U9PiPCNE8WarPYCFPD95NaBTPn7SyHC87fk3MQO+M96APMv+FP+Dv+gl4u/wDCfuv/AIxXueiXgu9LiKyX8rQgRPLeWcltJIwU
ZbY6r1znIGM5A6YrnB8RLeXwx4W1q2sfNj17UILIx+dt+zvIWDHO35trIw6DPXIru6ACiiigArPrQrPoAvJ9xfoKdTU+4v0FOoAK5/x3/wAk98S/9gq6/wDR
TV0FFAHz/rPjOzv/AIV/8Iva+TObfwtaXU1zDcB/LkWSKMwsoHDDOTk5Hp3q14jh08/D/SNP0XxbNq1y2t2Ttcz3qXklo7qVUYHRQVJCnvur1q38LaBbWlza
waHpkVtc7RcQJZxqk2Dkb1Aw2M8Z6UkHhTw9aIVtvD+lxIZElKx2cagumdjcDquTg9Rk460AeNxaVdyfDDx1eb2u/EllrjvNqDEBma1dG3r/AHQqGTA9yB1A
qK7hSf4V+FdS1C5l06bWfFa399dpL5TRmRph5qseEwiqQegxmvdE0jTUS8jXT7VY70s10ggUCcsMMXGPmyODnOajuNC0m9sYdPvNKsbiygx5NvNbo8ce0bV2
qRgYBIGOg4oA8zv9X0jRviL4Y1STWI7nSk0We2i1Ke5WRbh4zg5lHDOcc+pPqcVkw67aeJPiXoviDWdQn8MRy6AtwiDUFg83bePtjZ2C70dRu24GRj0zXrsn
hnQZtOh06XRNOexhYvFbNaRmKNjnJVcYBOT09TS33hjQNTkjlv8AQtNuXijESNcWschRBnCgkHAGTx05oA8K8d2811418YTSahLH4fF7pCaxFAgJkt2jGJA3
J+UgYAByXHpzY8XXr2198UrSBpvteo3Wl2dskIJeV2DEoAOTlFcY79O9e4zaDo9wt2s2kWMq3gjFyHtkInCfc35HzbcDGenakfQdGkumu30iwe5eZJ2ma3Qu
ZUyEctjO5cnB6jJxQBxHwi1FZH8V6ebO7sZItXe7W0uoDE8MM4DRrs/h4UnHvnvXn2g64z/E/S9cktNQaxu/EF+kWpNbsIbiOdFht41f+LaY2+XsB7HH0DFY
WtveXN3DaQR3Nzt8+ZIwry7RhdzDlsDgZ6VXj0LSI7S2s10qyS2tJRPbQrboEhkBJDouMK2STkc5JoA8E8FWU8/i/wALx317JPpc+rahPb2P3EhngXcJCR98
5Ix6H2JB66X/AI//AI0/9eKf+kklenLoWjwGAx6TYqYJWmh22yDy5G+868cMe5HJp/8AZOnFr5jp1puvhtuz5K/6QMFcScfOMEjnPBxQB49a3Hh2H4Oafpt5
fS65Z3klml+X1FXfSPNRRuBCny0jZOEYYzkEnJFZus+I7m9+HOsaFPd3mu2b6/DpdpfxYmlmiyJQFIwJXHl4z/EXHavarbwxoFnb3Nva6FpsFvdKFuIorSNV
mAzgOAMMBk9fWpI/DuiQwW1vHo1gkNrL59vGtsgWGTOd6DGFbPcc0AeCSpDr3gew8OSxXFjaJ42WzhtZYxHcW1tMHZQynOG/eP1z074qlbJfaR8PNZ0LVYLi
4mvPFMNjeiBTJPJgCRjEDjczGMYz13e9fRMmhaRNdtdyaTZPctNHO0zW6FzJGCEctjO5QSAeoycU+TSNNluVu3060kuVlEwlaBS4kChQ+7GdwUAZ64AFAHg9
jqkEXgzQLaSNrCw0fx1HDGl5iOSCDLy/vsn5WG9s9htPpXvthqFnqdol1p95BdWzkhZYJBIpwcHDAkHmqU3hnQLwT/atE02fz5RNN5tpG3mSAEBmyOWAJGTz
yfWrthp9npdotrYWdva2yElYYIxGgycnCgADmgC9RRRQAVn1oVn0AXk+4v0FOrPooA0KKz6KANCis+igDQorPooA0KKz6KANCis+igDQorPooA0KKz6KANCi
s+igDQorPooA0KKz6KANCis+igDQrPopKAP/2VBLAwQUAAYACAAAACEAPzTK8GACAABwBAAAGAAAAHhsL2RyYXdpbmdzL2RyYXdpbmcyLnhtbJxUW2+bMBR+
n7T/YPmdcCcpCqlGAlOlaqu6TXt2jQnewEa2m6at+t93bEKzizRNe+LDBx9/l2PWl8ehRwemNJeiwOEiwIgJKhsu9gX+8rn2VhhpQ0RDeilYgR+Zxpebt2/W
x0blD3qnEDQQOofXAnfGjLnva9qxgeiFHJmAaivVQAy8qr3fKPIArYfej4Ig8/WoGGl0x5jZTRV86kf+o9tAuMAbxwy4blnfvxO0k2paapUcJkRlvwnWvlVg
odsA4GPbbpYZ8Hot2RVXVfJh3mHhvGbr0SpdptMOKLkdrvP5OHY0iB4LnK7CIEoxoo8FzsKLELA/tRo5nYA43HB6cyJMPxxuFOJNgSOMBBnAfT6QPQsX38Y9
6CQ5dL7W5oTQveIFfq7rqEyrOvFqQF4SlIlXVsmFV0fxqoqW9TaKsxe7O8xyCu4bCP6qmV0Psz98HzhVUsvWLKgcfNm2nLI5R0gxTHznuyP6nKRVUKaryFsm
y62XxNvQK+Ok8rZVFZZ1epElu/LFyvYd+/npVEyBWNGnSABaNxAMScvULeuB7IHdMs2fwIvAtbEh/uLaXc/HmvcQK8ktPin7p+mcxO0kvR+YMNOIKnesFLrj
o8ZI5Wy4Y5CJumpCyBLuhgEyo+LCWEIk10YxQzsLW+Bxy6iZ9L4WnNAzTytBj1Y1yY+tgiElOTBBMDJuHDGCgYnjOFwlqRPtrPvLSIG3c59RafOeyQE81AZI
Axk3OeQAlk+05k/ssUJa6+z6mZSDrxNKew7W7IghqL2W9Lv+yk33yV7gnxP57fq5FvZnsfkBAAD//wMAUEsDBBQABgAIAAAAIQA5MbWR2wAAANABAAAjAAAA
eGwvd29ya3NoZWV0cy9fcmVscy9zaGVldDEueG1sLnJlbHOskc1qwzAMgO+DvoPRvXbSwxijTi9j0OvaPYBnK4lZIhtLW9e3n3coLKWwy276QZ8+oe3ua57U
JxaOiSy0ugGF5FOINFh4PT6vH0CxOApuSoQWzsiw61Z32xecnNQhHmNmVSnEFkaR/GgM+xFnxzplpNrpU5md1LQMJjv/7gY0m6a5N+U3A7oFU+2DhbIPG1DH
c66b/2anvo8en5L/mJHkxgoTijvVyyrSlQHFgtaXGl+CVldlMLdt2v+0ySWSYDmgSJXihdVVz1zlrX6L9CNpFn/ovgEAAP//AwBQSwMEFAAGAAgAAAAhAD50
UOPbAAAA0AEAACMAAAB4bC93b3Jrc2hlZXRzL19yZWxzL3NoZWV0Mi54bWwucmVsc6yRzWrDMAyA74O+g9G9dprDGKNOL2PQ69o9gGcriVkiG0tb17efdygs
pbDLbvpBnz6h7e5rntQnFo6JLGx0AwrJpxBpsPB6fF4/gGJxFNyUCC2ckWHXre62Lzg5qUM8xsyqUogtjCL50Rj2I86OdcpItdOnMjupaRlMdv7dDWjaprk3
5TcDugVT7YOFsg8tqOM5181/s1PfR49PyX/MSHJjhQnFneplFenKgGJB60uNL0GrqzKY2zab/7TJJZJgOaBIleKF1VXPXOWtfov0I2kWf+i+AQAA//8DAFBL
AwQUAAYACAAAACEAspc8br4AAAAkAQAAIwAAAHhsL2RyYXdpbmdzL19yZWxzL2RyYXdpbmcxLnhtbC5yZWxzhI/LigIxEEX3gv8Qam+q24UM0mk3MuB20A8o
kup0tPMgyQz69wbcjDAwy7qXew41HO5+ET+ci4tBQS87EBx0NC5YBZfz5+YDRKkUDC0xsIIHFziM69XwxQvVNiqzS0U0SigK5lrTHrHomT0VGROH1kwxe6rt
zBYT6RtZxm3X7TD/ZsD4xhQnoyCfTA/i/EjN/D87TpPTfIz623OofyjQ+eZuQMqWqwIp0bNx9Mp7eU0WcBzw7bfxCQAA//8DAFBLAwQUAAYACAAAACEAspc8
br4AAAAkAQAAIwAAAHhsL2RyYXdpbmdzL19yZWxzL2RyYXdpbmcyLnhtbC5yZWxzhI/LigIxEEX3gv8Qam+q24UM0mk3MuB20A8okup0tPMgyQz69wbcjDAw
y7qXew41HO5+ET+ci4tBQS87EBx0NC5YBZfz5+YDRKkUDC0xsIIHFziM69XwxQvVNiqzS0U0SigK5lrTHrHomT0VGROH1kwxe6rtzBYT6RtZxm3X7TD/ZsD4
xhQnoyCfTA/i/EjN/D87TpPTfIz623OofyjQ+eZuQMqWqwIp0bNx9Mp7eU0WcBzw7bfxCQAA//8DAFBLAwQUAAYACAAAACEAg6cIbcICAABcCgAAJwAAAHhs
L3ByaW50ZXJTZXR0aW5ncy9wcmludGVyU2V0dGluZ3MxLmJpbuxWzW7TQBD+nFBEBRI8QsWJS5qkTVvKLY1T5Co/JnYqjl2abWM1tS3baRoQEgeegmdAAgkO
PXDsgSMPwCtw5gjfOg4kadUfSFUO7Go969ndb74dz4xchgkLddQwhwqWsMIxR41EAIfPEGc37Ub65le8mk2/QyqFGby5XbjVgoa7aGspyraW5rOIwjk4l1nW
ks1KpjiU/MH22LDGzOhGrXkfW9qn9PFMavbbh7Ns3ElwBthDC1Mk/R/qn/fAZb76FjdbVXtDXeoeTO0F1pg7eSyz57HOeM/EUa+jxFkBi9CpzXBXKd6VoWaV
ukX2Fco8865IzRJeEtFw/W605rioCj4aZUuvVNB0nUCGamYKXwaW81yiUrbtcgNV2XKE3fclLLtY04sNHfVulGCsi23Z9LHhPTPFrqwHLRmgIQ9kEErUA0e6
kYgcz4VZb9iNomETntvi40+6ouNEfdS8YF90UPI6HRHxVC3eY21z1d3lqiuJGHqdbgykm8bicu6QQ53wgqrXkoPZxUPgeAbYLOjV4TfZNE3j85fvzQeEeMu1
Iw7VXlPOUqr8PeJwWbFySR+11kbE7uMRsuwhttFmdduH4Hye0qEmgMc3DzvcOc93j/oselxz0eJbj6tZLBA9j4ec5fhds8RUldLlGfXcHdOM2tmjvT4xPO5v
xUgGz/ok+ZHDp2bnFPdMl3eO0aZ4K/YX5b1DvoKekAn3vZjlb84LJ1hP29eKc55+/xvOiuX7xM+DG4zTvn4/T8bHIJ5VXOzx9vkJL1+Fj3OsPn8ezwe/WKo8
HPKejI7rj41JPyveA5Zt8j5kpHfOzcMec6HHGtFjPs3HGa3y/imq/IdSsbrKrnLMTTInpD8EK4qMkbdiOyGrwMk27p/T7agKpGrQ0KI1Us0Uop/gO1eAn4mr
XMhqJ3i7wY1U7VVNydH5TwAAAP//AwBQSwMEFAAGAAgAAAAhAPPu5+vCAgAAXAoAACcAAAB4bC9wcmludGVyU2V0dGluZ3MvcHJpbnRlclNldHRpbmdzMi5i
aW7sVs1u00AQ/pxQRAQSPELFiUuapE1byi2tU+QqPyZ2Ko5Zmm1jNbUt22laEBIHnoJnQAIJDj1w7IEjD8ArcOYI3zoOJGnVH0hVDuxqPevZ3W++Hc+MXIYJ
C3XUMIsKFrHMMUuNRACHzxBnN+1G+uZXvMqk3yGVQgZvbhdvtaHhLjpairKjpfksoXgOzmWWtWSzkikOJX+wPTasMTO6UWveR0v7lD6eSWW+fTjLxp0EZ4A9
tDBF0v+h/nkPXOart7jZqtob6lL3YGovsMrcKWCJvYB1xns2jnoda5wVsQCd2ix3rcW7stSsULfAvkxZYN6VqFnESyIart+LVh0XVcFHo2zplQqarhPIUM1M
4cvAcp5LVMq2XW6gKtuOsA99Ccsu1fRSQ0e9FyUY62JLNn1seM9MsSPrQVsGaMh9GYQS9cCRbiQix3Nh1ht2o2TYhOe2+PiTnug60SFqXrAnuljzul0R8VQt
3mNtcdXd4aoriRh63V4MpJvGwlL+gEOd8IKq15aD2cVD4HgG2Czq1eE32TRN4/OX780HhHjLtSMO1V5TZihV/h5xuKxY+aSPWusgYvfxCDn2EFvosLrtQXA+
R+lQE8Djm4dt7pzju0d9Dn2uuWjzrc/VHOaJXsBDzvL8rjliqkrp8ox67oxpRu3s0t4hMTzub8dIBs/6JPmRw6dm+xT3TJd3ntGmeCv2F+W9Tb6CnpAJ992Y
5W/O8ydYT9vXinOBfv8bzorl+8TPgxuM075+P0/GxyCeVVzs8vaFCS9fhY/zrD5/Hs/7v1iqPBzynoyO64+NST8r3gOWHfI+YKR3z83DPnOhzxrRZz7NxRmt
8v4pqvyHUrG6wq5yzE0yJ6Q/BCuKjJFbsZ2QVeBkG/fP6XZUBVI1aGjRGqlmCtFP8J0rwM/GVS5ktRO83eBGqvaqpuTo/CcAAAD//wMAUEsDBBQABgAIAAAA
IQB+xIpNVQEAAI4CAAARAAgBZG9jUHJvcHMvY29yZS54bWwgogQBKKAAAQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACMkl9PgzAUxd9N/A6k79AC28QGWHS6J5csOqPxrWnvNiIU0lbZvr0d/8Tpg4/tOffXc24azw9F7nyC0lkp
E+R7BDkgeSkyuUvQ82bpRsjRhknB8lJCgo6g0Ty9vIh5RXmpYK3KCpTJQDuWJDXlVYL2xlQUY833UDDtWYe04rZUBTP2qHa4Yvyd7QAHhMxwAYYJZhg+Ad1q
IKIOKfiArD5U3gAEx5BDAdJo7Hs+/vYaUIX+c6BRRs4iM8fKdurijtmCt+LgPuhsMNZ17dVhE8Pm9/Hr6uGpqepm8rQrDiiNBadcATOlSm8W948xHl2clpcz
bVZ2z9sMxO2x8/y+761rlUkDIg1IMHNJ5JJw41/TaUSnV28x7uZ6k3286domAOHY9LTt2isv4eJus0RnvElEQ2J5Z/OnNi2w6BL/jxgQSkI6mY2IPSBtQv/8
QekXAAAA//8DAFBLAwQUAAYACAAAACEAIOo6OaABAAA5AwAAEAAIAWRvY1Byb3BzL2FwcC54bWwgogQBKKAAAQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA
AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACck8Fu2zAMhu8D9g6G7oncdCiGQFYxpCty2LBgcXvYjZXpRJssCRJj
JHubPctebLK9uM56GNAbyZ/+9YmUxe2xMVmLIWpnC3Y1z1mGVrlK213BHsr72XuWRQJbgXEWC3bCyG7l2zdiE5zHQBpjlixsLNieyC85j2qPDcR5km1Sahca
oJSGHXd1rRXeOXVo0BJf5PkNxyOhrbCa+dGQDY7Lll5rWjnV8cXH8uQTsBQfvDdaAaVbys9aBRddTdnHo0Ij+FQUiW6L6hA0nWQu+DQVWwUGV8lY1mAiCv5c
EGuEbmgb0CFK0dKyRUUuZFH/TGNbsOwJInY4BWshaLCUsLq2Ielj4yMFuXbfIWYVZur3L6MOxgme+gatD6efTGP9Ti76hhRcNnYGA08SLklLTQbjl3oDgf4H
3jMM2APO/az8+m2W58OxfxFH2LN6/eIC/WgSyj+Hr1zjwZ6SMEaftP0RH3zp7oDwPPbLotjuIWCVNjWuZSyIdZp4MJ3Jag92h9W556XQPZLH4U+QVzfz/DpP
+5/UBH9+8/IPAAAA//8DAFBLAQItABQABgAIAAAAIQBrK/rSlgEAAK8GAAATAAAAAAAAAAAAAAAAAAAAAABbQ29udGVudF9UeXBlc10ueG1sUEsBAi0AFAAG
AAgAAAAhALVVMCP0AAAATAIAAAsAAAAAAAAAAAAAAAAAzwMAAF9yZWxzLy5yZWxzUEsBAi0AFAAGAAgAAAAhAMOzx5coAwAALwcAAA8AAAAAAAAAAAAAAAAA
9AYAAHhsL3dvcmtib29rLnhtbFBLAQItABQABgAIAAAAIQBKqaZh+gAAAEcDAAAaAAAAAAAAAAAAAAAAAEkKAAB4bC9fcmVscy93b3JrYm9vay54bWwucmVs
c1BLAQItABQABgAIAAAAIQD/BbKFeQoAANc2AAAYAAAAAAAAAAAAAAAAAIMMAAB4bC93b3Jrc2hlZXRzL3NoZWV0MS54bWxQSwECLQAUAAYACAAAACEAhHUx
NRoJAACJKAAAGAAAAAAAAAAAAAAAAAAyFwAAeGwvd29ya3NoZWV0cy9zaGVldDIueG1sUEsBAi0AFAAGAAgAAAAhAINNbMhWBwAAyCAAABMAAAAAAAAAAAAA
AAAAgiAAAHhsL3RoZW1lL3RoZW1lMS54bWxQSwECLQAUAAYACAAAACEAlwwfs04FAACxKgAADQAAAAAAAAAAAAAAAAAJKAAAeGwvc3R5bGVzLnhtbFBLAQIt
ABQABgAIAAAAIQABtcQrCwIAAM0FAAAUAAAAAAAAAAAAAAAAAIItAAB4bC9zaGFyZWRTdHJpbmdzLnhtbFBLAQItABQABgAIAAAAIQC1fjtOWgIAAG0EAAAY
AAAAAAAAAAAAAAAAAL8vAAB4bC9kcmF3aW5ncy9kcmF3aW5nMS54bWxQSwECLQAKAAAAAAAAACEAzKVD4ywfAAAsHwAAEwAAAAAAAAAAAAAAAABPMgAAeGwv
bWVkaWEvaW1hZ2UxLmpwZ1BLAQItABQABgAIAAAAIQA/NMrwYAIAAHAEAAAYAAAAAAAAAAAAAAAAAKxRAAB4bC9kcmF3aW5ncy9kcmF3aW5nMi54bWxQSwEC
LQAUAAYACAAAACEAOTG1kdsAAADQAQAAIwAAAAAAAAAAAAAAAABCVAAAeGwvd29ya3NoZWV0cy9fcmVscy9zaGVldDEueG1sLnJlbHNQSwECLQAUAAYACAAA
ACEAPnRQ49sAAADQAQAAIwAAAAAAAAAAAAAAAABeVQAAeGwvd29ya3NoZWV0cy9fcmVscy9zaGVldDIueG1sLnJlbHNQSwECLQAUAAYACAAAACEAspc8br4A
AAAkAQAAIwAAAAAAAAAAAAAAAAB6VgAAeGwvZHJhd2luZ3MvX3JlbHMvZHJhd2luZzEueG1sLnJlbHNQSwECLQAUAAYACAAAACEAspc8br4AAAAkAQAAIwAA
AAAAAAAAAAAAAAB5VwAAeGwvZHJhd2luZ3MvX3JlbHMvZHJhd2luZzIueG1sLnJlbHNQSwECLQAUAAYACAAAACEAg6cIbcICAABcCgAAJwAAAAAAAAAAAAAA
AAB4WAAAeGwvcHJpbnRlclNldHRpbmdzL3ByaW50ZXJTZXR0aW5nczEuYmluUEsBAi0AFAAGAAgAAAAhAPPu5+vCAgAAXAoAACcAAAAAAAAAAAAAAAAAf1sA
AHhsL3ByaW50ZXJTZXR0aW5ncy9wcmludGVyU2V0dGluZ3MyLmJpblBLAQItABQABgAIAAAAIQB+xIpNVQEAAI4CAAARAAAAAAAAAAAAAAAAAIZeAABkb2NQ
cm9wcy9jb3JlLnhtbFBLAQItABQABgAIAAAAIQAg6jo5oAEAADkDAAAQAAAAAAAAAAAAAAAAABJhAABkb2NQcm9wcy9hcHAueG1sUEsFBgAAAAAUABQAgQUA
AOhjAAAAAA==
"""


def ahora_colombia():
    return datetime.now(ZONA_HORARIA_COLOMBIA)


def limpiar_observacion_operacion(valor):
    texto = "" if valor is None else str(valor).strip()
    texto = re.sub(r"\s*\[OP:[^\]]+\]\s*", " ", texto)
    return " ".join(texto.split())


def capacidad_barril_desde_codigo(codigo):
    codigo = normalizar_codigo_barril(codigo)
    return {"20": 20, "30": 30, "58": 58}.get(codigo[:2])


def nombre_producto_barril(codigo, estilo):
    capacidad = capacidad_barril_desde_codigo(codigo)
    partes = ["Barril"]
    if capacidad:
        partes.append(f"{capacidad} L")
    if str(estilo).strip():
        partes.append(str(estilo).strip())
    return " - ".join(partes)


def nombre_producto_latas(estilo):
    estilo = str(estilo).strip()
    return f"Latas 330 mL - {estilo}" if estilo else "Latas 330 mL"


def mapa_direcciones_clientes_fresco():
    try:
        clientes = cargar_csv_publico_fresco(HOJA_CLIENTES)
    except Exception:
        return {}
    if clientes.empty or "Nombre" not in clientes.columns or "Dirección" not in clientes.columns:
        return {}
    mapa = {}
    for _, fila in clientes.iterrows():
        nombre = str(fila.get("Nombre", "") or "").strip()
        direccion = str(fila.get("Dirección", "") or "").strip()
        if nombre:
            mapa[normalizar_clave(nombre)] = direccion
    return mapa


def registrar_pedido_local_orden_general(
    barriles, latas, cliente, responsable, observaciones, direccion=""
):
    """Mantiene visible el pedido recién guardado mientras Google Sheets termina de actualizar."""
    fecha = ahora_colombia().date().isoformat()
    registros = list(st.session_state.get("orden_general_registros_locales", []))
    for barril in barriles:
        codigo = normalizar_codigo_barril(barril.get("codigo", ""))
        registros.append({
            "Fecha": fecha,
            "Cliente": str(cliente).strip(),
            "Dirección": str(direccion or "").strip(),
            "Producto": nombre_producto_barril(codigo, barril.get("estilo", "")),
            "Cantidad": 1,
            "Lote": str(barril.get("lote", "") or "").strip(),
            "Código del Barril": codigo,
            "Responsable": str(responsable).strip(),
            "Observaciones": limpiar_observacion_operacion(observaciones),
            "Tipo": "Barril",
        })
    for lata in latas:
        registros.append({
            "Fecha": fecha,
            "Cliente": str(cliente).strip(),
            "Dirección": str(direccion or "").strip(),
            "Producto": nombre_producto_latas(lata.get("estilo", "")),
            "Cantidad": int(lata.get("cantidad", 0) or 0),
            "Lote": str(lata.get("lote", "") or "").strip(),
            "Código del Barril": "",
            "Responsable": str(responsable).strip(),
            "Observaciones": limpiar_observacion_operacion(observaciones),
            "Tipo": "Latas",
        })
    st.session_state.orden_general_registros_locales = registros[-500:]
    st.session_state.pop("documentos_orden_general", None)


def _pedidos_barriles_sheet(fecha_obj, direcciones):
    datos = cargar_datos_barriles_frescos()
    if datos.empty:
        return []
    fechas = datos["Marca temporal"]
    mascara = (
        fechas.notna()
        & fechas.dt.date.eq(fecha_obj)
        & datos["_estado_key"].eq("despacho")
    )
    filas = []
    for _, fila in datos.loc[mascara].iterrows():
        codigo = normalizar_codigo_barril(fila.get("Código", ""))
        cliente = str(fila.get("Cliente", "") or "").strip()
        if not codigo or not cliente:
            continue
        estilo = str(fila.get("Estilo", "") or "").strip()
        filas.append({
            "Fecha": fecha_obj.isoformat(),
            "Cliente": cliente,
            "Dirección": direcciones.get(normalizar_clave(cliente), ""),
            "Producto": nombre_producto_barril(codigo, estilo),
            "Cantidad": 1,
            "Lote": str(fila.get("Lote", "") or "").strip(),
            "Código del Barril": codigo,
            "Responsable": str(fila.get("Responsable", "") or "").strip(),
            "Observaciones": limpiar_observacion_operacion(fila.get("Observaciones", "")),
            "Tipo": "Barril",
        })
    return filas


def _pedidos_latas_sheet(fecha_obj, direcciones):
    try:
        datos = cargar_csv_publico_fresco(HOJA_MOVIMIENTOS_LATAS)
    except Exception:
        return []
    if datos.empty:
        return []
    requeridas = {"Marca temporal", "Estilo", "Cantidad", "Lote", "Cliente"}
    if not requeridas.issubset(datos.columns):
        return []

    fechas = convertir_fechas_sheet(datos["Marca temporal"])
    estados = (
        datos["Estado"].map(normalizar_clave)
        if "Estado" in datos.columns
        else pd.Series("", index=datos.index)
    )
    cantidades = limpiar_cantidades(datos["Cantidad"])
    mascara = fechas.notna() & fechas.dt.date.eq(fecha_obj) & estados.isin(["", "despacho"])
    filas = []
    for indice in datos.index[mascara]:
        cliente = str(datos.at[indice, "Cliente"] or "").strip()
        cantidad = int(cantidades.at[indice])
        if not cliente or cantidad <= 0:
            continue
        estilo = str(datos.at[indice, "Estilo"] or "").strip()
        responsable = (
            str(datos.at[indice, "Responsable"] or "").strip()
            if "Responsable" in datos.columns else ""
        )
        observaciones = (
            limpiar_observacion_operacion(datos.at[indice, "Observaciones"])
            if "Observaciones" in datos.columns else ""
        )
        filas.append({
            "Fecha": fecha_obj.isoformat(),
            "Cliente": cliente,
            "Dirección": direcciones.get(normalizar_clave(cliente), ""),
            "Producto": nombre_producto_latas(estilo),
            "Cantidad": cantidad,
            "Lote": str(datos.at[indice, "Lote"] or "").strip(),
            "Código del Barril": "",
            "Responsable": responsable,
            "Observaciones": observaciones,
            "Tipo": "Latas",
        })
    return filas


def combinar_pedidos_sheet_y_locales(filas_sheet, fecha_obj):
    columnas = [
        "Fecha", "Cliente", "Dirección", "Producto", "Cantidad", "Lote",
        "Código del Barril", "Responsable", "Observaciones", "Tipo"
    ]
    locales = [
        fila for fila in st.session_state.get("orden_general_registros_locales", [])
        if str(fila.get("Fecha", "")) == fecha_obj.isoformat()
    ]
    sheet_df = pd.DataFrame(filas_sheet, columns=columnas)
    local_df = pd.DataFrame(locales, columns=columnas)

    barriles_sheet = sheet_df[sheet_df["Tipo"].eq("Barril")].copy() if not sheet_df.empty else pd.DataFrame(columns=columnas)
    barriles_local = local_df[local_df["Tipo"].eq("Barril")].copy() if not local_df.empty else pd.DataFrame(columns=columnas)
    barriles = pd.concat([barriles_sheet, barriles_local], ignore_index=True)
    if not barriles.empty:
        barriles["_cliente_key"] = barriles["Cliente"].map(normalizar_clave)
        barriles["_codigo_key"] = barriles["Código del Barril"].map(normalizar_codigo_barril)
        barriles.drop_duplicates(["_cliente_key", "_codigo_key"], keep="first", inplace=True)
        barriles.drop(columns=["_cliente_key", "_codigo_key"], inplace=True)

    def agrupar_latas(df):
        if df.empty:
            return pd.DataFrame(columns=columnas)
        latas_df = df[df["Tipo"].eq("Latas")].copy()
        if latas_df.empty:
            return pd.DataFrame(columns=columnas)
        latas_df["Cantidad"] = pd.to_numeric(latas_df["Cantidad"], errors="coerce").fillna(0).astype(int)
        claves = ["Fecha", "Cliente", "Dirección", "Producto", "Lote", "Código del Barril", "Tipo"]
        agrupado = (
            latas_df.groupby(claves, dropna=False, as_index=False)
            .agg({
                "Cantidad": "sum",
                "Responsable": lambda s: " / ".join(dict.fromkeys(str(v).strip() for v in s if str(v).strip())),
                "Observaciones": lambda s: " | ".join(dict.fromkeys(str(v).strip() for v in s if str(v).strip())),
            })
        )
        return agrupado[columnas]

    latas_sheet = agrupar_latas(sheet_df)
    latas_local = agrupar_latas(local_df)
    if latas_sheet.empty:
        latas = latas_local
    elif latas_local.empty:
        latas = latas_sheet
    else:
        claves = ["Fecha", "Cliente", "Dirección", "Producto", "Lote", "Código del Barril", "Tipo"]
        unidos = latas_sheet.merge(
            latas_local,
            on=claves,
            how="outer",
            suffixes=("_sheet", "_local"),
        )
        unidos["Cantidad"] = unidos[["Cantidad_sheet", "Cantidad_local"]].max(axis=1).fillna(0).astype(int)
        for campo in ("Responsable", "Observaciones"):
            unidos[campo] = unidos[f"{campo}_sheet"].fillna("")
            faltan = unidos[campo].astype(str).str.strip().eq("")
            unidos.loc[faltan, campo] = unidos.loc[faltan, f"{campo}_local"].fillna("")
        latas = unidos[claves + ["Cantidad", "Responsable", "Observaciones"]]
        latas = latas[columnas]

    resultado = pd.concat([barriles, latas], ignore_index=True)
    if resultado.empty:
        return resultado
    resultado["Cantidad"] = pd.to_numeric(resultado["Cantidad"], errors="coerce").fillna(0).astype(int)
    resultado["_cliente_key"] = resultado["Cliente"].map(normalizar_clave)
    resultado["_tipo_orden"] = resultado["Tipo"].map({"Barril": 0, "Latas": 1}).fillna(9)
    resultado.sort_values(
        ["_cliente_key", "_tipo_orden", "Producto", "Lote", "Código del Barril"],
        inplace=True,
        kind="stable",
    )
    resultado.drop(columns=["_cliente_key", "_tipo_orden"], inplace=True)
    resultado.reset_index(drop=True, inplace=True)
    return resultado


def cargar_orden_general_del_dia(fecha_obj):
    direcciones = mapa_direcciones_clientes_fresco()
    filas = _pedidos_barriles_sheet(fecha_obj, direcciones)
    filas.extend(_pedidos_latas_sheet(fecha_obj, direcciones))
    return combinar_pedidos_sheet_y_locales(filas, fecha_obj)


def localizar_plantilla_formatos():
    """Localiza la plantilla externa o reconstruye la copia incorporada en la app."""
    bases = []
    try:
        bases.append(Path(__file__).resolve().parent)
    except NameError:
        pass
    bases.extend([Path.cwd(), Path("/mnt/data")])

    # Primero respeta cualquier plantilla externa existente, por si se reemplaza
    # posteriormente por una versión oficial más reciente.
    for base in bases:
        for nombre in NOMBRES_PLANTILLA_FORMATOS:
            candidato = base / nombre
            if candidato.is_file():
                return candidato

    # Si no existe un archivo externo, reconstruye la plantilla interna.
    try:
        contenido = base64.b64decode(
            "".join(PLANTILLA_FORMATOS_EMBEBIDA_BASE64.split()),
            validate=True,
        )
        if not contenido.startswith(b"PK"):
            raise ValueError("La copia interna no tiene formato XLSX válido.")

        destino = Path("/tmp") / "FORMATOS_DE_DESPACHOS_CASTIZA.xlsx"
        if not destino.exists() or destino.stat().st_size != len(contenido):
            destino.write_bytes(contenido)
        return destino
    except Exception as exc:
        raise FileNotFoundError(
            "No fue posible cargar la plantilla externa ni reconstruir la copia interna "
            f"de los formatos: {exc}"
        ) from exc


def combinar_textos_unicos(valores, separador=" | "):
    salida = []
    vistos = set()
    for valor in valores:
        texto = str(valor or "").strip()
        clave = normalizar_clave(texto)
        if texto and clave not in vistos:
            vistos.add(clave)
            salida.append(texto)
    return separador.join(salida)


def construir_bloques_clientes(pedidos_df):
    bloques = []
    if pedidos_df is None or pedidos_df.empty:
        return bloques
    grupos = []
    for cliente, grupo in pedidos_df.groupby("Cliente", sort=False, dropna=False):
        grupos.append((str(cliente), grupo.reset_index(drop=True)))
    grupos.sort(key=lambda par: normalizar_clave(par[0]))

    for cliente, grupo in grupos:
        direccion = combinar_textos_unicos(grupo["Dirección"].tolist(), " / ")
        observaciones = combinar_textos_unicos(grupo["Observaciones"].tolist())
        responsables = combinar_textos_unicos(grupo["Responsable"].tolist(), " / ")
        items = grupo.to_dict("records")
        for inicio in range(0, len(items), 3):
            bloques.append({
                "cliente": cliente,
                "direccion": direccion,
                "observaciones": observaciones,
                "responsables": responsables,
                "continuacion": inicio > 0,
                "items": items[inicio:inicio + 3],
            })
    return bloques


def dividir_paginas(lista, tamano=9):
    if not lista:
        return [[]]
    return [lista[i:i + tamano] for i in range(0, len(lista), tamano)]


def leer_logo_bytes_plantilla(ruta_plantilla=None):
    ruta = Path(ruta_plantilla or localizar_plantilla_formatos())
    try:
        with zipfile.ZipFile(ruta, "r") as archivo:
            imagenes = sorted(
                nombre for nombre in archivo.namelist()
                if nombre.startswith("xl/media/")
                and nombre.lower().endswith((".png", ".jpg", ".jpeg", ".gif"))
            )
            return archivo.read(imagenes[0]) if imagenes else b""
    except Exception:
        return b""


def copiar_hoja_con_imagenes(libro, origen, titulo, logo_bytes=b""):
    destino = libro.copy_worksheet(origen)
    destino.title = titulo
    imagenes_origen = list(getattr(origen, "_images", []))
    if logo_bytes and imagenes_origen:
        referencia = imagenes_origen[0]
        buffer_imagen = io.BytesIO(logo_bytes)
        nueva = XLImage(buffer_imagen)
        nueva._buffer_orden_general = buffer_imagen
        nueva.width = referencia.width
        nueva.height = referencia.height
        nueva.anchor = copy(referencia.anchor)
        destino.add_image(nueva)
    return destino


def preparar_hojas_paginas(libro, nombre_base, total_paginas, logo_bytes=b""):
    origen = libro[nombre_base]
    hojas = []
    origen.title = nombre_base if total_paginas == 1 else f"{nombre_base} P1"
    hojas.append(origen)
    for numero in range(2, total_paginas + 1):
        hojas.append(
            copiar_hoja_con_imagenes(
                libro, origen, f"{nombre_base} P{numero}", logo_bytes
            )
        )
    return hojas


def configurar_impresion_hoja(hoja, area):
    hoja.sheet_view.showGridLines = False
    hoja.print_area = area
    hoja.page_setup.orientation = "landscape"
    hoja.page_setup.fitToWidth = 1
    hoja.page_setup.fitToHeight = 1
    hoja.sheet_properties.pageSetUpPr.fitToPage = True
    hoja.print_options.horizontalCentered = True
    hoja.print_options.verticalCentered = True


def obtener_celda_escribible(hoja, coordenada):
    """Devuelve la celda real; si pertenece a un rango combinado usa su esquina superior izquierda."""
    for rango in hoja.merged_cells.ranges:
        if coordenada in rango:
            return hoja.cell(row=rango.min_row, column=rango.min_col)
    return hoja[coordenada]


def asignar_valor_celda(hoja, coordenada, valor):
    """Escribe de forma segura incluso cuando la coordenada pertenece a celdas combinadas."""
    celda = obtener_celda_escribible(hoja, coordenada)
    celda.value = valor
    return celda


def escribir_texto_celda(hoja, coordenada, valor, horizontal=None, tamano=None, negrita=None):
    celda = asignar_valor_celda(hoja, coordenada, valor)
    alineacion = copy(celda.alignment)
    celda.alignment = Alignment(
        horizontal=horizontal or alineacion.horizontal,
        vertical=alineacion.vertical or "center",
        text_rotation=alineacion.text_rotation,
        wrap_text=True,
        shrink_to_fit=alineacion.shrink_to_fit,
        indent=alineacion.indent,
    )
    if tamano is not None or negrita is not None:
        fuente = copy(celda.font)
        celda.font = Font(
            name=fuente.name,
            sz=tamano if tamano is not None else fuente.sz,
            bold=negrita if negrita is not None else fuente.bold,
            italic=fuente.italic,
            vertAlign=fuente.vertAlign,
            underline=fuente.underline,
            strike=fuente.strike,
            color=fuente.color,
        )


def limpiar_bloques_formato(hoja, ultima_columna):
    for indice in range(9):
        inicio = 7 + indice * 3
        for columna in range(1, ultima_columna + 1):
            for fila in range(inicio, inicio + 3):
                celda = hoja.cell(fila, columna)
                if celda.__class__.__name__ != "MergedCell":
                    celda.value = None


def llenar_formato_002(
    hoja, bloques, fecha_obj, pagina, total_paginas,
    conductor, realiza, supervisa, calidad_cumple, limpieza_cumple
):
    limpiar_bloques_formato(hoja, 11)
    asignar_valor_celda(hoja, "K4", f"Página {pagina} de {total_paginas}")
    fecha_texto = fecha_obj.strftime("%d/%m/%Y")
    for indice, bloque in enumerate(bloques):
        inicio = 7 + indice * 3
        cliente = bloque["cliente"] + (" (continuación)" if bloque["continuacion"] else "")
        escribir_texto_celda(hoja, f"A{inicio}", fecha_texto, "center", 9)
        escribir_texto_celda(hoja, f"B{inicio}", cliente, "center", 8)
        escribir_texto_celda(hoja, f"K{inicio}", bloque["observaciones"], "center", 8)
        asignar_valor_celda(hoja, f"G{inicio}", "X" if calidad_cumple else "")
        asignar_valor_celda(hoja, f"H{inicio}", "" if calidad_cumple else "X")
        asignar_valor_celda(hoja, f"I{inicio}", "X" if limpieza_cumple else "")
        asignar_valor_celda(hoja, f"J{inicio}", "" if limpieza_cumple else "X")
        for desplazamiento, item in enumerate(bloque["items"]):
            fila = inicio + desplazamiento
            escribir_texto_celda(hoja, f"C{fila}", item.get("Producto", ""), "left", 8)
            escribir_texto_celda(hoja, f"D{fila}", int(item.get("Cantidad", 0) or 0), "center", 9)
            escribir_texto_celda(hoja, f"E{fila}", item.get("Lote", ""), "center", 8)
            escribir_texto_celda(hoja, f"F{fila}", item.get("Código del Barril", ""), "center", 9)

    escribir_texto_celda(hoja, "C34", conductor, "left", 9)
    escribir_texto_celda(hoja, "B35", realiza, "left", 9)
    escribir_texto_celda(hoja, "B36", supervisa, "left", 9)
    configurar_impresion_hoja(hoja, "A1:K36")


def llenar_formato_003(
    hoja, bloques, fecha_obj, pagina, total_paginas, realiza, supervisa
):
    limpiar_bloques_formato(hoja, 8)
    asignar_valor_celda(hoja, "H4", f"Página {pagina} de {total_paginas}")
    fecha_texto = fecha_obj.strftime("%d/%m/%Y")
    for indice, bloque in enumerate(bloques):
        inicio = 7 + indice * 3
        cliente = bloque["cliente"] + (" (continuación)" if bloque["continuacion"] else "")
        cliente_direccion = cliente + ("\n" + bloque["direccion"] if bloque["direccion"] else "")
        escribir_texto_celda(hoja, f"A{inicio}", fecha_texto, "center", 9)
        escribir_texto_celda(hoja, f"B{inicio}", cliente_direccion, "center", 8)
        escribir_texto_celda(hoja, f"G{inicio}", bloque["observaciones"], "center", 8)
        for desplazamiento, item in enumerate(bloque["items"]):
            fila = inicio + desplazamiento
            escribir_texto_celda(hoja, f"C{fila}", item.get("Producto", ""), "left", 8)
            escribir_texto_celda(hoja, f"D{fila}", int(item.get("Cantidad", 0) or 0), "center", 9)

        # Estas casillas están combinadas verticalmente para todo el bloque del cliente.
        # Se escribe solo en la esquina superior izquierda del rango combinado.
        asignar_valor_celda(hoja, f"E{inicio}", "")
        asignar_valor_celda(hoja, f"F{inicio}", "")
        asignar_valor_celda(hoja, f"H{inicio}", "")

    escribir_texto_celda(hoja, "B34", realiza, "left", 9)
    escribir_texto_celda(hoja, "B35", supervisa, "left", 9)
    configurar_impresion_hoja(hoja, "A1:H35")


def generar_archivo_formatos_orden_general(
    pedidos_df, fecha_obj, conductor="", realiza="", supervisa="",
    calidad_cumple=True, limpieza_cumple=True
):
    bloques = construir_bloques_clientes(pedidos_df)
    paginas = dividir_paginas(bloques, 9)
    ruta_plantilla = localizar_plantilla_formatos()
    logo_bytes = leer_logo_bytes_plantilla(ruta_plantilla)
    libro = load_workbook(ruta_plantilla)

    hojas_002 = preparar_hojas_paginas(
        libro, "F-TRZ-002", len(paginas), logo_bytes
    )
    hojas_003 = preparar_hojas_paginas(
        libro, "F-TRZ-003", len(paginas), logo_bytes
    )
    libro._sheets = hojas_002 + hojas_003

    for numero, bloques_pagina in enumerate(paginas, start=1):
        llenar_formato_002(
            hojas_002[numero - 1], bloques_pagina, fecha_obj, numero, len(paginas),
            conductor, realiza, supervisa, calidad_cumple, limpieza_cumple
        )
        llenar_formato_003(
            hojas_003[numero - 1], bloques_pagina, fecha_obj, numero, len(paginas),
            realiza, supervisa
        )

    libro.active = 0
    salida = io.BytesIO()
    libro.save(salida)
    salida.seek(0)
    return salida.getvalue(), bloques, paginas


def extraer_logo_plantilla_base64():
    datos = leer_logo_bytes_plantilla()
    return base64.b64encode(datos).decode("ascii") if datos else ""


def _html_items_bloque(bloque):
    items = list(bloque.get("items", []))
    while len(items) < 3:
        items.append({})
    return items[:3]


def generar_html_impresion_orden_general(
    pedidos_df, fecha_obj, conductor="", realiza="", supervisa="",
    calidad_cumple=True, limpieza_cumple=True
):
    bloques = construir_bloques_clientes(pedidos_df)
    paginas = dividir_paginas(bloques, 9)
    logo = extraer_logo_plantilla_base64()
    logo_html = f'<img class="logo" src="data:image/png;base64,{logo}">' if logo else '<div class="logo-text">CASTIZA</div>'
    fecha_texto = fecha_obj.strftime("%d/%m/%Y")

    def e(valor):
        return html_lib.escape(str(valor or ""))

    paginas_html = []
    total = len(paginas)
    for numero, bloques_pagina in enumerate(paginas, start=1):
        filas_002 = []
        for bloque in bloques_pagina:
            items = _html_items_bloque(bloque)
            cliente = bloque["cliente"] + (" (continuación)" if bloque["continuacion"] else "")
            for indice, item in enumerate(items):
                celdas = []
                if indice == 0:
                    celdas.extend([
                        f'<td rowspan="3">{e(fecha_texto)}</td>',
                        f'<td rowspan="3">{e(cliente)}</td>',
                    ])
                celdas.extend([
                    f'<td class="left">{e(item.get("Producto", ""))}</td>',
                    f'<td>{e(item.get("Cantidad", ""))}</td>',
                    f'<td>{e(item.get("Lote", ""))}</td>',
                    f'<td>{e(item.get("Código del Barril", ""))}</td>',
                ])
                if indice == 0:
                    celdas.extend([
                        f'<td rowspan="3">{"X" if calidad_cumple else ""}</td>',
                        f'<td rowspan="3">{"" if calidad_cumple else "X"}</td>',
                        f'<td rowspan="3">{"X" if limpieza_cumple else ""}</td>',
                        f'<td rowspan="3">{"" if limpieza_cumple else "X"}</td>',
                        f'<td rowspan="3" class="small">{e(bloque["observaciones"])}</td>',
                    ])
                filas_002.append("<tr>" + "".join(celdas) + "</tr>")
        for _ in range(9 - len(bloques_pagina)):
            for indice in range(3):
                celdas = []
                if indice == 0:
                    celdas.extend(['<td rowspan="3"></td>', '<td rowspan="3"></td>'])
                celdas.extend(['<td></td>', '<td></td>', '<td></td>', '<td></td>'])
                if indice == 0:
                    celdas.extend([
                        '<td rowspan="3"></td>', '<td rowspan="3"></td>',
                        '<td rowspan="3"></td>', '<td rowspan="3"></td>',
                        '<td rowspan="3"></td>'
                    ])
                filas_002.append("<tr>" + "".join(celdas) + "</tr>")
        paginas_html.append(f'''
        <section class="page legal">
          <header>{logo_html}<div class="title">Formato de Pedidos y distribución de producto terminado</div>
          <div class="meta">Programa P-TRZ-10<br>Código: F-TRZ-002<br>Versión: 01<br>Página {numero} de {total}</div></header>
          <table><thead><tr><th rowspan="2">Fecha</th><th rowspan="2">Nombre del Cliente</th><th rowspan="2">Producto</th><th rowspan="2">Cantidad</th><th rowspan="2">Lote</th><th rowspan="2">Código del Barril</th><th colspan="2">Calidad del producto</th><th colspan="2">Limpieza del vehículo</th><th rowspan="2">Observaciones</th></tr>
          <tr><th>C</th><th>NC</th><th>C</th><th>NC</th></tr></thead><tbody>{''.join(filas_002)}</tbody></table>
          <div class="footer-grid"><div><b>Responsable del transporte:</b> {e(conductor)}</div><div><b>Realiza:</b> {e(realiza)}</div><div><b>Supervisa:</b> {e(supervisa)}</div><div><b>Convenciones:</b> C: Cumple · NC: No cumple</div></div>
        </section>''')

    for numero, bloques_pagina in enumerate(paginas, start=1):
        filas_003 = []
        for bloque in bloques_pagina:
            items = _html_items_bloque(bloque)
            cliente = bloque["cliente"] + (" (continuación)" if bloque["continuacion"] else "")
            cliente_dir = e(cliente) + ("<br>" + e(bloque["direccion"]) if bloque["direccion"] else "")
            for indice, item in enumerate(items):
                celdas = []
                if indice == 0:
                    celdas.extend([
                        f'<td rowspan="3">{e(fecha_texto)}</td>',
                        f'<td rowspan="3">{cliente_dir}</td>',
                    ])
                celdas.extend([
                    f'<td class="left">{e(item.get("Producto", ""))}</td>',
                    f'<td>{e(item.get("Cantidad", ""))}</td>',
                ])
                if indice == 0:
                    celdas.extend([
                        '<td rowspan="3"></td>', '<td rowspan="3"></td>',
                        f'<td rowspan="3" class="small">{e(bloque["observaciones"])}</td>',
                        '<td rowspan="3"></td>',
                    ])
                filas_003.append("<tr>" + "".join(celdas) + "</tr>")
        for _ in range(9 - len(bloques_pagina)):
            for indice in range(3):
                celdas = []
                if indice == 0:
                    celdas.extend(['<td rowspan="3"></td>', '<td rowspan="3"></td>'])
                celdas.extend(['<td></td>', '<td></td>'])
                if indice == 0:
                    celdas.extend([
                        '<td rowspan="3"></td>', '<td rowspan="3"></td>',
                        '<td rowspan="3"></td>', '<td rowspan="3"></td>'
                    ])
                filas_003.append("<tr>" + "".join(celdas) + "</tr>")
        paginas_html.append(f'''
        <section class="page a4">
          <header>{logo_html}<div class="title">Formato de orden de entrega de producto</div>
          <div class="meta">Programa P-TRZ-10<br>Código: F-TRZ-003<br>Versión: 001<br>Página {numero} de {total}</div></header>
          <table><thead><tr><th rowspan="2">Fecha</th><th rowspan="2">Nombre del Cliente y Dirección</th><th rowspan="2">Producto</th><th rowspan="2">Cantidad</th><th colspan="2">Pedido Correcto</th><th rowspan="2">Observaciones</th><th rowspan="2">Firma de recibido</th></tr>
          <tr><th>SI</th><th>NO</th></tr></thead><tbody>{''.join(filas_003)}</tbody></table>
          <div class="footer-grid"><div><b>Realiza:</b> {e(realiza)}</div><div><b>Supervisa:</b> {e(supervisa)}</div></div>
        </section>''')

    return f'''<!doctype html><html><head><meta charset="utf-8"><style>
      @page {{ size: landscape; margin: 7mm; }}
      body {{ font-family: Arial, sans-serif; color:#111; margin:0; background:#eee; }}
      .toolbar {{ position:sticky; top:0; z-index:10; padding:10px; text-align:center; background:#fff; border-bottom:1px solid #bbb; }}
      .toolbar button {{ padding:10px 18px; font-size:16px; font-weight:bold; cursor:pointer; }}
      .page {{ background:#fff; width:calc(100% - 24px); margin:12px auto; padding:8mm; box-sizing:border-box; page-break-after:always; }}
      header {{ display:grid; grid-template-columns:140px 1fr 170px; align-items:center; border:1px solid #111; min-height:66px; }}
      .logo {{ max-width:115px; max-height:62px; margin:auto; }} .logo-text {{font-weight:bold;text-align:center;}}
      .title {{ text-align:center; font-weight:bold; font-size:17px; }} .meta {{ border-left:1px solid #111; padding:5px; font-size:11px; line-height:1.35; }}
      table {{ width:100%; border-collapse:collapse; table-layout:fixed; font-size:9px; }}
      th, td {{ border:1px solid #111; padding:3px; text-align:center; vertical-align:middle; height:18px; overflow-wrap:anywhere; }}
      th {{ font-weight:bold; }} td.left {{ text-align:left; }} td.small {{ font-size:8px; }}
      .footer-grid {{ display:grid; grid-template-columns:repeat(2,1fr); gap:8px 20px; margin-top:8px; font-size:10px; }}
      @media print {{ body {{ background:#fff; }} .toolbar {{ display:none; }} .page {{ margin:0; width:100%; box-shadow:none; }} }}
    </style></head><body><div class="toolbar"><button onclick="window.print()">🖨️ Imprimir los dos formatos</button></div>{''.join(paginas_html)}</body></html>'''

def solicitar_guardado_movimiento():
    """Callback: el primer clic bloquea inmediatamente cualquier clic posterior."""
    if not st.session_state.get("movimiento_guardando", False):
        st.session_state.movimiento_guardando = True
        st.session_state.movimiento_solicitado = True


def guardar_resultado_y_reiniciar(tipo, mensaje, detalles=None, conservar_operacion=False):
    st.session_state.resultado_movimiento = {
        "tipo": tipo,
        "mensaje": mensaje,
        "detalles": detalles or [],
    }
    st.session_state.movimiento_guardando = False
    st.session_state.movimiento_solicitado = False
    if not conservar_operacion:
        limpiar_operacion_pendiente()
    st.rerun()


def limpiar_widgets_movimiento():
    claves = [
        "mov_codigo_despacho", "mov_codigo_manual", "mov_lote", "mov_estilo",
        "mov_observaciones", "incluye_latas_despacho"
    ]
    cantidad_ordenes = int(st.session_state.get("num_latas", 1))
    for indice in range(cantidad_ordenes):
        claves.extend(
            [f"estilo_lata_{indice}", f"lote_lata_{indice}", f"cantidad_lata_{indice}"]
        )
    for clave in claves:
        st.session_state.pop(clave, None)
    st.session_state.num_latas = 1
    st.session_state.barriles_pedido_despacho = []

# --- Lista de estilos global ---
estilos = ["Golden", "Amber", "Vienna Lager", "Brown Ale Cafe", "Stout",
           "Session IPA", "IPA", "Maracuyá", "Barley Wine", "Trigo", "Catharina Sour",
           "Gose", "Imperial IPA", "NEIPA", "Imperial Stout", "Otros"]

# IMAGEN DE FONDO PERSONALIZADA Y ESTILOS GENERALES
if os.path.exists("background.jpg"):
    with open("background.jpg", "rb") as img:
        encoded = base64.b64encode(img.read()).decode()
    st.markdown(
        f"""
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Roboto&display=swap');
        html, body, [class*="st"]  {{
            font-family: 'Roboto', sans-serif;
            color: #fff3aa;
        }}
        .stApp {{
            background-image: url("data:image/jpeg;base64,{encoded}");
            background-size: cover;
            background-position: center;
            background-repeat: no-repeat;
            background-attachment: fixed;
        }}
        .stTextInput > div > div > input,
        .stSelectbox > div > div,
        .stTextArea > div > textarea {{
            background-color: #ffffff10 !important;
            color: #fff3aa !important;
            border-radius: 10px;
        }}
        .stButton > button {{
            background-color: #55dcad !important;
            color: #fff3aa !important;
            border: none;
            border-radius: 10px;
            font-weight: bold;
        }}
        .stDataFrame, .stTable {{
            background-color: rgba(0,0,0,0.6);
            border-radius: 10px;
        }}
        </style>
        """,
        unsafe_allow_html=True
    )

# TÍTULO PRINCIPAL
st.markdown("<h1 style='text-align:center; color:#fff3aa;'>🍺 Sistema de Trazabilidad de Barriles y Latas - Castiza</h1>", unsafe_allow_html=True)

import streamlit as st
import pandas as pd
import requests

# ---------- TÍTULO ----------
st.markdown("<h2 style='color:#fff3aa;'>📋🛢️ Registro Movimiento Barriles</h2>", unsafe_allow_html=True)

# ---------- RESULTADO DEL ÚLTIMO INTENTO ----------
resultado_anterior = st.session_state.pop("resultado_movimiento", None)
if resultado_anterior:
    tipo = resultado_anterior.get("tipo", "info")
    mensaje = resultado_anterior.get("mensaje", "")

    if tipo == "success":
        # Celebración visual sin mostrar el resumen ni el detalle de cada barril.
        st.balloons()
        st.markdown(
            """
            <style>
            .beer-celebration {
                position: fixed;
                inset: 0;
                pointer-events: none;
                overflow: hidden;
                z-index: 999999;
            }
            .beer-celebration span {
                position: absolute;
                top: -12vh;
                font-size: 2.2rem;
                animation-name: beer-fall;
                animation-timing-function: linear;
                animation-fill-mode: forwards;
            }
            @keyframes beer-fall {
                0%   { transform: translateY(-12vh) rotate(0deg); opacity: 0; }
                10%  { opacity: 1; }
                100% { transform: translateY(115vh) rotate(540deg); opacity: 0.95; }
            }
            </style>
            <div class="beer-celebration" aria-hidden="true">
                <span style="left:5%;animation-duration:3.7s;animation-delay:0.0s;">🍺</span>
                <span style="left:14%;animation-duration:4.4s;animation-delay:0.4s;">🍻</span>
                <span style="left:24%;animation-duration:3.9s;animation-delay:0.2s;">🍺</span>
                <span style="left:34%;animation-duration:4.8s;animation-delay:0.7s;">🍻</span>
                <span style="left:45%;animation-duration:3.6s;animation-delay:0.1s;">🍺</span>
                <span style="left:56%;animation-duration:4.3s;animation-delay:0.5s;">🍻</span>
                <span style="left:66%;animation-duration:3.8s;animation-delay:0.3s;">🍺</span>
                <span style="left:76%;animation-duration:4.6s;animation-delay:0.8s;">🍻</span>
                <span style="left:86%;animation-duration:4.0s;animation-delay:0.2s;">🍺</span>
                <span style="left:94%;animation-duration:4.5s;animation-delay:0.6s;">🍻</span>
            </div>
            """,
            unsafe_allow_html=True,
        )
    elif tipo == "warning":
        st.warning(f"⚠️ {mensaje}")
        for detalle in resultado_anterior.get("detalles", []):
            st.write(f"• {detalle}")
    else:
        st.error(f"❌ {mensaje}")
        for detalle in resultado_anterior.get("detalles", []):
            st.write(f"• {detalle}")

if BACKEND_SEGURO_ACTIVO:
    st.success("🔒 Protección transaccional activa: bloqueo simultáneo e idempotencia habilitados.")

# ---------- ESTADO Y DATOS ACTUALES DE BARRILES ----------
if "movimiento_guardando" not in st.session_state:
    st.session_state.movimiento_guardando = False
if "movimiento_solicitado" not in st.session_state:
    st.session_state.movimiento_solicitado = False
if "barriles_pedido_despacho" not in st.session_state:
    st.session_state.barriles_pedido_despacho = []

estado_barril = st.selectbox(
    "Estado del barril",
    ["Despacho", "Lavado en bodega", "Sucio", "En cuarto frío"],
    key="mov_estado",
    disabled=st.session_state.movimiento_guardando,
)

try:
    datos_barriles_ui = cargar_datos_barriles_ui()
    ultimos_barriles_ui = obtener_ultimos_movimientos(datos_barriles_ui)
except Exception as exc:
    datos_barriles_ui = pd.DataFrame()
    ultimos_barriles_ui = pd.DataFrame()
    st.error(f"No fue posible consultar el estado actual de los barriles: {exc}")

codigo_barril = ""
lote_producto = ""
estilo_cerveza = ""
ultimo_ui = None

if estado_barril == "Despacho":
    cuarto_frio_ui = barriles_actualmente_en_cuarto_frio(datos_barriles_ui)
    cuarto_frio_ui.sort_values(["Estilo", "Código"], inplace=True, ignore_index=True)
    mapa_barriles = {
        fila["Código"]: {
            "codigo": fila["Código"],
            "estilo": fila["Estilo"],
            "lote": fila["Lote"],
        }
        for _, fila in cuarto_frio_ui.iterrows()
    }

    barriles_pedido = list(st.session_state.get("barriles_pedido_despacho", []))
    codigos_en_pedido = {
        normalizar_codigo_barril(item.get("codigo", "")) for item in barriles_pedido
    }
    codigos_disponibles = [
        codigo for codigo in mapa_barriles if codigo not in codigos_en_pedido
    ]
    opciones = [""] + codigos_disponibles
    if st.session_state.get("mov_codigo_despacho", "") not in opciones:
        st.session_state.mov_codigo_despacho = ""

    codigo_para_agregar = st.selectbox(
        "Selecciona un barril para agregar al pedido",
        opciones,
        key="mov_codigo_despacho",
        format_func=lambda codigo: (
            "— Selecciona un barril —"
            if codigo == ""
            else (
                f"{codigo} — {mapa_barriles[codigo]['estilo'] or 'Sin estilo'}"
                f" — lote {mapa_barriles[codigo]['lote'] or 'sin lote'}"
            )
        ),
        disabled=st.session_state.movimiento_guardando,
        help="La lista contiene únicamente barriles cuyo último estado es En cuarto frío.",
    )

    if st.button(
        "➕ Agregar barril al pedido",
        key="agregar_barril_pedido",
        use_container_width=True,
        disabled=st.session_state.movimiento_guardando or not codigo_para_agregar,
    ):
        datos_agregar = dict(mapa_barriles[codigo_para_agregar])
        st.session_state.barriles_pedido_despacho = barriles_pedido + [datos_agregar]
        st.rerun()

    barriles_pedido = list(st.session_state.get("barriles_pedido_despacho", []))
    st.markdown("#### 🧾 Lista de barriles del pedido")
    if barriles_pedido:
        tabla_pedido = pd.DataFrame([
            {
                "Código": item.get("codigo", ""),
                "Estilo": item.get("estilo", ""),
                "Lote": item.get("lote", ""),
            }
            for item in barriles_pedido
        ])
        st.dataframe(tabla_pedido, hide_index=True, use_container_width=True)
        st.caption(f"Total de barriles seleccionados: {len(barriles_pedido)}")

        for indice, item in enumerate(barriles_pedido):
            col_info, col_quitar = st.columns([7, 1])
            col_info.write(
                f"**{indice + 1}. {item.get('codigo', '')}** · "
                f"{item.get('estilo', '') or 'Sin estilo'} · "
                f"lote {item.get('lote', '') or 'sin lote'}"
            )
            if col_quitar.button(
                "🗑️",
                key=f"quitar_barril_pedido_{indice}_{item.get('codigo', '')}",
                disabled=st.session_state.movimiento_guardando,
                help="Quitar este barril del pedido.",
            ):
                nueva_lista = list(barriles_pedido)
                nueva_lista.pop(indice)
                st.session_state.barriles_pedido_despacho = nueva_lista
                st.rerun()
    else:
        st.info(
            "Todavía no has agregado barriles. Puedes agregar varios o continuar con un despacho exclusivo de latas."
        )
        if cuarto_frio_ui.empty:
            st.caption("No hay barriles disponibles actualmente en cuarto frío.")

else:
    codigo_barril = st.text_input(
        "Código del barril (5 dígitos; inicia por 20, 30 o 58)",
        key="mov_codigo_manual",
        disabled=st.session_state.movimiento_guardando,
    ).strip()
    if codigo_barril:
        ultimo_ui = obtener_ultimo_movimiento(datos_barriles_ui, codigo_barril)
        if ultimo_ui is not None:
            st.caption(
                f"Último estado registrado: {ultimo_ui['Estado'] or 'Sin estado'}"
                + (f" · cliente: {ultimo_ui['Cliente']}" if ultimo_ui["Cliente"] else "")
            )
            if normalizar_clave(ultimo_ui["Estado"]) == normalizar_clave(estado_barril):
                st.error(f"🚫 Este barril ya tiene como último estado '{estado_barril}'.")
        else:
            st.caption("El código no tiene movimientos anteriores registrados.")

if estado_barril == "En cuarto frío":
    lote_producto = st.text_input(
        "Lote del producto",
        key="mov_lote",
        disabled=st.session_state.movimiento_guardando,
    ).strip()
    estilo_cerveza = st.selectbox(
        "Estilo",
        estilos,
        key="mov_estilo",
        disabled=st.session_state.movimiento_guardando,
    )

# ---------- CLIENTES ----------
try:
    df_clientes = cargar_hoja_csv(HOJA_CLIENTES).copy()
    df_clientes.columns = df_clientes.columns.astype(str).str.strip()
    if "Nombre" not in df_clientes.columns:
        raise ValueError("La hoja RClientes no contiene la columna Nombre.")
    df_clientes["Nombre"] = df_clientes["Nombre"].fillna("").astype(str).str.strip()
    df_clientes = df_clientes[df_clientes["Nombre"].ne("")]
    lista_clientes = df_clientes["Nombre"].drop_duplicates().tolist()
    if "Dirección" in df_clientes.columns:
        dict_direcciones = df_clientes.drop_duplicates("Nombre").set_index("Nombre")["Dirección"].to_dict()
    else:
        dict_direcciones = {}
except Exception as exc:
    lista_clientes = []
    dict_direcciones = {}
    if estado_barril == "Despacho":
        st.warning(f"No se pudieron cargar los clientes: {exc}")

cliente = "Planta Castiza"
if estado_barril == "Despacho":
    if lista_clientes:
        cliente = st.selectbox(
            "Cliente",
            lista_clientes,
            key="mov_cliente",
            disabled=st.session_state.movimiento_guardando,
        )
        direccion_cliente = dict_direcciones.get(cliente, "")
        if direccion_cliente:
            st.text_input(
                "Dirección del cliente",
                value=str(direccion_cliente),
                disabled=True,
                key="mov_direccion_cliente",
            )
    else:
        cliente = st.text_input(
            "Cliente",
            key="mov_cliente_manual",
            disabled=st.session_state.movimiento_guardando,
        ).strip()

# ---------- DESPACHO DE LATAS ----------
latas = []
errores_configuracion_latas = []
inventario_latas = pd.DataFrame()
incluye_latas = "No"

if estado_barril == "Despacho":
    incluye_latas = st.selectbox(
        "¿🚚 Incluye despacho de latas?",
        ["No", "Sí"],
        key="incluye_latas_despacho",
        disabled=st.session_state.movimiento_guardando,
    )

    if incluye_latas == "Sí":
        st.markdown(
            "<h4 style='color:#fff3aa;'>🚚 Detalles de despacho de latas</h4>",
            unsafe_allow_html=True,
        )

        if st.button(
            "🔄 Actualizar inventario de latas",
            key="actualizar_inventario_latas",
            disabled=st.session_state.movimiento_guardando,
        ):
            cargar_hoja_csv.clear()
            st.rerun()

        try:
            inventario_latas = calcular_inventario_latas()
        except Exception as exc:
            st.error(f"❌ No se pudo calcular el inventario de latas: {exc}")
            errores_configuracion_latas.append("No fue posible consultar el inventario de latas.")

        if inventario_latas.empty:
            st.warning("⚠️ No hay lotes de latas con existencias disponibles.")
            errores_configuracion_latas.append("No hay inventario de latas disponible para despachar.")
        else:
            resumen_estilos = (
                inventario_latas.groupby("Estilo", as_index=False)["Disponible"]
                .sum()
                .sort_values("Estilo", key=lambda serie: serie.astype(str).str.casefold())
            )
            st.caption("Existencias actuales por estilo")
            st.dataframe(resumen_estilos, hide_index=True, use_container_width=True)

            if "num_latas" not in st.session_state:
                st.session_state.num_latas = 1

            estilos_disponibles = inventario_latas["Estilo"].drop_duplicates().tolist()
            opcion_vacia_estilo = "— Selecciona un estilo —"
            opcion_vacia_lote = "— Selecciona un lote —"

            for indice in range(st.session_state.num_latas):
                st.markdown(f"### Orden {indice + 1}")
                estilo_lata = st.selectbox(
                    f"Estilo de cerveza (Orden {indice + 1})",
                    [opcion_vacia_estilo] + estilos_disponibles,
                    key=f"estilo_lata_{indice}",
                    disabled=st.session_state.movimiento_guardando,
                )

                if estilo_lata == opcion_vacia_estilo:
                    errores_configuracion_latas.append(
                        f"Selecciona el estilo de la orden {indice + 1}."
                    )
                    st.selectbox(
                        f"Lote disponible (Orden {indice + 1})",
                        [opcion_vacia_lote],
                        key=f"lote_lata_{indice}",
                        disabled=True,
                    )
                    continue

                inventario_estilo = inventario_latas[
                    inventario_latas["Estilo"] == estilo_lata
                ].copy()
                disponibilidad_por_lote = {
                    fila.Lote: int(fila.Disponible)
                    for fila in inventario_estilo.itertuples(index=False)
                }
                lotes_disponibles = list(disponibilidad_por_lote.keys())
                lote_lata = st.selectbox(
                    f"Lote disponible (Orden {indice + 1})",
                    [opcion_vacia_lote] + lotes_disponibles,
                    format_func=lambda lote, mapa=disponibilidad_por_lote: (
                        lote if lote == opcion_vacia_lote
                        else f"{lote} — {mapa[lote]} latas disponibles"
                    ),
                    key=f"lote_lata_{indice}",
                    disabled=st.session_state.movimiento_guardando,
                )

                if lote_lata == opcion_vacia_lote:
                    errores_configuracion_latas.append(
                        f"Selecciona el lote de la orden {indice + 1}."
                    )
                    continue

                disponible_lote = disponibilidad_por_lote[lote_lata]
                cantidad_key = f"cantidad_lata_{indice}"
                if cantidad_key in st.session_state:
                    cantidad_actual = int(st.session_state[cantidad_key])
                    st.session_state[cantidad_key] = min(max(cantidad_actual, 1), disponible_lote)

                cantidad = st.number_input(
                    f"Cantidad (Orden {indice + 1})",
                    min_value=1,
                    max_value=disponible_lote,
                    step=1,
                    key=cantidad_key,
                    disabled=st.session_state.movimiento_guardando,
                    help=f"Máximo disponible para este lote: {disponible_lote}",
                )
                latas.append(
                    {"cantidad": int(cantidad), "estilo": estilo_lata, "lote": lote_lata}
                )

            columnas_orden = st.columns(2)
            if columnas_orden[0].button(
                "➕ Agregar otro lote",
                key="agregar_orden_latas",
                disabled=st.session_state.movimiento_guardando,
            ):
                st.session_state.num_latas += 1
                st.rerun()

            if columnas_orden[1].button(
                "➖ Quitar última orden",
                key="quitar_orden_latas",
                disabled=(
                    st.session_state.movimiento_guardando
                    or st.session_state.num_latas <= 1
                ),
            ):
                indice = st.session_state.num_latas - 1
                for prefijo in ("estilo_lata_", "lote_lata_", "cantidad_lata_"):
                    st.session_state.pop(f"{prefijo}{indice}", None)
                st.session_state.num_latas -= 1
                st.rerun()

# ---------- RESPONSABLE Y OBSERVACIONES ----------
responsables = [
    "Pepe Vallejo", "Ligia Cajigas", "Erika Martinez", "Marcelo Martinez",
    "Yimer Luna", "Operario 2"
]
responsable = st.selectbox(
    "Responsable",
    responsables,
    key="mov_responsable",
    disabled=st.session_state.movimiento_guardando,
)
observaciones = st.text_area(
    "Observaciones",
    key="mov_observaciones",
    disabled=st.session_state.movimiento_guardando,
)

st.caption(
    "Al presionar Guardar, el botón se bloquea hasta terminar. No es necesario volver a hacer clic aunque la conexión esté lenta."
)
st.button(
    "🔒 Guardar Registro",
    key="guardar_movimiento_principal",
    type="primary",
    use_container_width=True,
    disabled=st.session_state.movimiento_guardando,
    on_click=solicitar_guardado_movimiento,
)

# ---------- PROCESAR EXACTAMENTE UNA VEZ POR CLIC ----------
if st.session_state.get("movimiento_solicitado", False):
    errores_guardado = []
    barriles_despacho = list(st.session_state.get("barriles_pedido_despacho", []))
    registrar_barril = bool(codigo_barril.strip()) if estado_barril != "Despacho" else bool(barriles_despacho)
    despacho_solo_latas = (
        estado_barril == "Despacho"
        and incluye_latas == "Sí"
        and not barriles_despacho
    )

    if estado_barril == "Despacho":
        if not barriles_despacho and not despacho_solo_latas:
            errores_guardado.append(
                "Debes agregar al menos un barril a la lista o seleccionar un despacho exclusivo de latas."
            )
    elif not registrar_barril:
        errores_guardado.append("Debes ingresar un código de barril.")

    if estado_barril == "Despacho" and not str(cliente).strip():
        errores_guardado.append("Debes seleccionar o ingresar el cliente del despacho.")

    datos_barriles_frescos = pd.DataFrame()
    barriles_validados = []
    ultimo_fresco = None

    if estado_barril == "Despacho" and barriles_despacho:
        try:
            cargar_datos_barriles_ui.clear()
            datos_barriles_frescos = cargar_datos_barriles_frescos()
        except Exception as exc:
            errores_guardado.append(
                f"No se pudo verificar el estado actual de los barriles. No se registró nada: {exc}"
            )
        else:
            codigos_repetidos = set()
            vistos = set()
            for item in barriles_despacho:
                codigo_item = normalizar_codigo_barril(item.get("codigo", ""))
                if codigo_item in vistos:
                    codigos_repetidos.add(codigo_item)
                vistos.add(codigo_item)
            if codigos_repetidos:
                errores_guardado.append(
                    "Hay barriles repetidos en el pedido: " + ", ".join(sorted(codigos_repetidos))
                )

            for item in barriles_despacho:
                codigo_item = normalizar_codigo_barril(item.get("codigo", ""))
                valido, mensaje, ultimo_item = validar_movimiento_barril_local(
                    datos_barriles_frescos,
                    codigo_item,
                    "Despacho",
                    estilo=item.get("estilo", ""),
                    lote=item.get("lote", ""),
                )
                if not valido:
                    errores_guardado.append(f"Barril {codigo_item}: {mensaje}")
                    continue
                if movimiento_local_reciente(codigo_item, "Despacho"):
                    errores_guardado.append(
                        f"Barril {codigo_item}: este despacho fue confirmado recientemente y se bloqueó para evitar un duplicado."
                    )
                    continue
                barriles_validados.append({
                    "codigo": codigo_item,
                    "estilo": str(ultimo_item.get("Estilo", "")).strip() if ultimo_item is not None else str(item.get("estilo", "")).strip(),
                    "lote": str(ultimo_item.get("Lote", "")).strip() if ultimo_item is not None else str(item.get("lote", "")).strip(),
                })

    elif registrar_barril:
        try:
            cargar_datos_barriles_ui.clear()
            datos_barriles_frescos = cargar_datos_barriles_frescos()
        except Exception as exc:
            errores_guardado.append(
                f"No se pudo verificar el estado actual del barril. No se registró nada: {exc}"
            )
        else:
            valido, mensaje, ultimo_fresco = validar_movimiento_barril_local(
                datos_barriles_frescos,
                codigo_barril,
                estado_barril,
                estilo=estilo_cerveza,
                lote=lote_producto,
            )
            if not valido:
                errores_guardado.append(mensaje)

            if movimiento_local_reciente(codigo_barril, estado_barril):
                errores_guardado.append(
                    "Este mismo movimiento fue confirmado recientemente en esta sesión. "
                    "Se bloqueó para evitar un duplicado mientras Google actualiza la hoja."
                )

            if estado_barril == "Despacho" and ultimo_fresco is not None:
                estilo_cerveza = str(ultimo_fresco.get("Estilo", "")).strip()
                lote_producto = str(ultimo_fresco.get("Lote", "")).strip()

    inventario_actual = inventario_latas
    if incluye_latas == "Sí":
        errores_guardado.extend(errores_configuracion_latas)
        if not errores_configuracion_latas:
            try:
                cargar_hoja_csv.clear()
                inventario_actual = calcular_inventario_latas()
                errores_guardado.extend(validar_existencias_latas(latas, inventario_actual))
            except Exception as exc:
                errores_guardado.append(
                    f"No fue posible validar el inventario actualizado de latas: {exc}"
                )

    if errores_guardado:
        guardar_resultado_y_reiniciar(
            "error",
            "La operación fue bloqueada y no se registró ningún movimiento.",
            list(dict.fromkeys(errores_guardado)),
            conservar_operacion=False,
        )

    if estado_barril == "Despacho":
        datos_huella = {
            "barriles": barriles_validados,
            "estado": "Despacho",
            "cliente": cliente,
            "responsable": responsable,
            "observaciones": observaciones,
            "incluye_latas": incluye_latas,
            "latas": latas,
        }
    else:
        datos_huella = {
            "codigo": codigo_barril,
            "estado": estado_barril,
            "cliente": cliente,
            "responsable": responsable,
            "observaciones": observaciones,
            "lote": lote_producto,
            "estilo": estilo_cerveza,
            "incluye_latas": incluye_latas,
            "latas": latas,
        }

    huella = construir_huella_operacion(datos_huella)
    operacion_id = obtener_operacion_id_para_huella(huella)

    with st.spinner("Validando y registrando. No cierres la página ni vuelvas a presionar el botón..."):
        if estado_barril == "Despacho":
            exito, mensaje, reintentable, respuesta_extra = registrar_lista_despacho(
                barriles_validados,
                {
                    "cliente": cliente,
                    "responsable": responsable,
                    "observaciones": observaciones,
                    "latas": latas if incluye_latas == "Sí" else [],
                },
                operacion_id,
            )
        else:
            payload_seguro = dict(datos_huella)
            payload_seguro["operacion_id"] = operacion_id
            if BACKEND_SEGURO_ACTIVO:
                exito, mensaje, reintentable, respuesta_extra = enviar_backend_apps_script(payload_seguro)
            else:
                exito, mensaje, reintentable, respuesta_extra = enviar_por_formularios_compatibilidad(payload_seguro)

    if exito:
        if estado_barril == "Despacho":
            for barril in barriles_validados:
                marcar_movimiento_local_confirmado(
                    barril["codigo"], "Despacho", crear_id_suboperacion(
                        operacion_id, "barril", barril["codigo"]
                    )
                )
        else:
            marcar_movimiento_local_confirmado(codigo_barril, estado_barril, operacion_id)

        if estado_barril == "Despacho":
            direccion_orden = ""
            try:
                direccion_orden = str(dict_direcciones.get(cliente, "") or "").strip()
            except Exception:
                direccion_orden = ""
            registrar_pedido_local_orden_general(
                barriles_validados,
                latas if incluye_latas == "Sí" else [],
                cliente,
                responsable,
                observaciones,
                direccion_orden,
            )

        cargar_hoja_csv.clear()
        cargar_datos_barriles_ui.clear()
        limpiar_widgets_movimiento()
        detalles = list(respuesta_extra.get("detalles", []))
        guardar_resultado_y_reiniciar(
            "success",
            mensaje,
            detalles,
            conservar_operacion=False,
        )
    else:
        tipo_resultado = "warning" if reintentable or respuesta_extra.get("partial") else "error"
        detalles = list(respuesta_extra.get("detalles", []))
        if reintentable:
            detalles.append(
                f"Identificador conservado: {operacion_id}. Al reintentar, los barriles ya registrados no se duplicarán."
            )
        guardar_resultado_y_reiniciar(
            tipo_resultado,
            mensaje,
            detalles,
            conservar_operacion=reintentable or bool(respuesta_extra.get("partial")),
        )

# ---------- ORDEN GENERAL DEL DÍA ----------
st.markdown("---")
st.markdown(
    "<h2 style='color:#fff3aa;'>📦 Orden General</h2>",
    unsafe_allow_html=True,
)
fecha_orden_general = ahora_colombia().date()
st.caption(
    "Reúne los despachos registrados hoy por cliente y prepara los formatos "
    "F-TRZ-002 y F-TRZ-003."
)

col_fecha_orden, col_actualizar_orden = st.columns([3, 1])
col_fecha_orden.info(f"Fecha de la orden: {fecha_orden_general.strftime('%d/%m/%Y')}")
if col_actualizar_orden.button(
    "🔄 Actualizar",
    key="actualizar_orden_general_dia",
    use_container_width=True,
):
    cargar_hoja_csv.clear()
    cargar_datos_barriles_ui.clear()
    st.session_state.pop("documentos_orden_general", None)
    st.rerun()

try:
    pedidos_orden_general = cargar_orden_general_del_dia(fecha_orden_general)
except Exception as exc:
    pedidos_orden_general = pd.DataFrame()
    st.error(f"No fue posible construir la orden general del día: {exc}")

if pedidos_orden_general.empty:
    st.info("Todavía no hay pedidos de despacho registrados para la fecha actual.")
else:
    tabla_visible = pedidos_orden_general[[
        "Cliente", "Dirección", "Producto", "Cantidad", "Lote",
        "Código del Barril", "Responsable", "Observaciones"
    ]].copy()
    st.dataframe(tabla_visible, hide_index=True, use_container_width=True)

    total_clientes = pedidos_orden_general["Cliente"].nunique()
    total_barriles = int((pedidos_orden_general["Tipo"] == "Barril").sum())
    total_latas = int(
        pedidos_orden_general.loc[
            pedidos_orden_general["Tipo"] == "Latas", "Cantidad"
        ].sum()
    )
    metrica_clientes, metrica_barriles, metrica_latas = st.columns(3)
    metrica_clientes.metric("Clientes", total_clientes)
    metrica_barriles.metric("Barriles", total_barriles)
    metrica_latas.metric("Latas", total_latas)

    responsables_dia = [
        valor for valor in pedidos_orden_general["Responsable"].dropna().astype(str).unique()
        if valor.strip()
    ]
    if "orden_general_realiza" not in st.session_state:
        st.session_state.orden_general_realiza = responsables_dia[0] if len(responsables_dia) == 1 else ""

    st.markdown("#### Datos para completar los formatos")
    conductor_orden = st.text_input(
        "Responsable del transporte / conductor",
        key="orden_general_conductor",
    )
    col_realiza, col_supervisa = st.columns(2)
    realiza_orden = col_realiza.text_input(
        "Realiza",
        key="orden_general_realiza",
    )
    supervisa_orden = col_supervisa.text_input(
        "Supervisa",
        key="orden_general_supervisa",
    )
    col_calidad, col_limpieza = st.columns(2)
    calidad_cumple = col_calidad.checkbox(
        "Calidad del producto: Cumple",
        value=True,
        key="orden_general_calidad_cumple",
    )
    limpieza_cumple = col_limpieza.checkbox(
        "Limpieza del vehículo: Cumple",
        value=True,
        key="orden_general_limpieza_cumple",
    )

    if st.button(
        "📄 Preparar formatos de despacho y entrega",
        type="primary",
        use_container_width=True,
        key="preparar_documentos_orden_general",
    ):
        try:
            archivo_excel, bloques_excel, paginas_excel = generar_archivo_formatos_orden_general(
                pedidos_orden_general,
                fecha_orden_general,
                conductor=conductor_orden,
                realiza=realiza_orden,
                supervisa=supervisa_orden,
                calidad_cumple=calidad_cumple,
                limpieza_cumple=limpieza_cumple,
            )
            html_impresion = generar_html_impresion_orden_general(
                pedidos_orden_general,
                fecha_orden_general,
                conductor=conductor_orden,
                realiza=realiza_orden,
                supervisa=supervisa_orden,
                calidad_cumple=calidad_cumple,
                limpieza_cumple=limpieza_cumple,
            )
            st.session_state.documentos_orden_general = {
                "excel": archivo_excel,
                "html": html_impresion,
                "nombre": f"orden_general_castiza_{fecha_orden_general.strftime('%Y%m%d')}.xlsx",
                "clientes": total_clientes,
                "paginas": len(paginas_excel),
                "lineas": len(pedidos_orden_general),
            }
            st.success(
                f"Formatos preparados: {total_clientes} cliente(s), "
                f"{len(pedidos_orden_general)} línea(s) y {len(paginas_excel)} página(s) por formato."
            )
        except Exception as exc:
            st.session_state.pop("documentos_orden_general", None)
            st.error(f"No se pudieron preparar los formatos: {exc}")

    documentos_orden = st.session_state.get("documentos_orden_general")
    if documentos_orden:
        st.download_button(
            "⬇️ Descargar F-TRZ-002 y F-TRZ-003",
            data=documentos_orden["excel"],
            file_name=documentos_orden["nombre"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="descargar_documentos_orden_general",
        )
        st.caption(
            "El archivo conserva ambos formatos y crea páginas adicionales automáticamente "
            "cuando hay más clientes o productos de los que caben en una sola hoja."
        )
        import streamlit.components.v1 as components
        with st.expander("🖨️ Imprimir los formatos", expanded=True):
            components.html(
                documentos_orden["html"],
                height=850,
                scrolling=True,
            )

# Mantener la variable utilizada por las secciones inferiores de búsqueda y últimos movimientos.
url_datos = URL_DATOS_BARRILES_PUBLICA

# FORMULARIO NUEVO CLIENTE
st.markdown("---")
st.markdown("<h2 style='color:#fff3aa;'>➕ Registrar Nuevo Cliente</h2>", unsafe_allow_html=True)

nuevo_cliente = st.text_input("Nombre del nuevo cliente")
direccion_nuevo_cliente = st.text_input("Dirección del cliente")

if st.button("Agregar Cliente"):
    if nuevo_cliente.strip() != "":
        form_cliente_url = "https://docs.google.com/forms/d/e/1FAIpQLScllMMM33p5F--_I6Y80gsLUsusGMTk0OA3XDVC9ocngoc2Hw/formResponse"
        payload_cliente = {
            "entry.1250409245": nuevo_cliente,
            "entry.82359015": direccion_nuevo_cliente
        }
        response = requests.post(form_cliente_url, data=payload_cliente)
        if response.status_code in [200, 302]:
            st.success("✅ Cliente agregado correctamente")
        else:
            st.error(f"❌ Error al agregar cliente. Código: {response.status_code}")

# ----------------------------------------
# ÚLTIMOS MOVIMIENTOS
# ----------------------------------------
st.markdown("---")
st.markdown("<h2 style='color:#fff3aa;'>📑 Últimos 10 Movimientos</h2>", unsafe_allow_html=True)

try:
    url_datos = "https://docs.google.com/spreadsheets/d/1FjQ8XBDwDdrlJZsNkQ6YyaygkHLhpKmfLBv6wd3uluY/gviz/tq?tqx=out:csv&sheet=DatosM"
    df_mov = pd.read_csv(url_datos)
    df_mov.columns = df_mov.columns.str.strip()
    if not df_mov.empty:
        df_mov = df_mov[df_mov["Código"].notna()]
        st.dataframe(df_mov.tail(10)[["Código", "Estilo", "Estado", "Cliente", "Responsable", "Observaciones"]])
    else:
        st.warning("⚠️ La hoja está vacía.")
except Exception as e:
    st.error(f"⚠️ No se pudo cargar la hoja de movimientos: {e}")

# FORMULARIO PARA INGRESAR LATAS AL CUARTO FRÍO
st.markdown("---")
st.markdown("<h2 style='color:#fff3aa;'>❄️ Ingreso de Latas al Cuarto Frío</h2>", unsafe_allow_html=True)
estilo_lata_cf = st.selectbox("Estilo de las latas", estilos, key="estilo_cf")
cantidad_lata_cf = st.number_input("Cantidad de latas", min_value=1, key="cantidad_cf")
lote_lata_cf = st.text_input("Lote", key="lote_cf")

if st.button("Guardar Ingreso de Latas al Cuarto Frío"):
    # ✅ URL CORRECTA para enviar POST
    form_cf_url = "https://docs.google.com/forms/d/e/1FAIpQLScWQqIe6atpchC8I4nxFRxzLa17AgQfA4v1v86Gtz_3nYg60Q/formResponse"

    # ✅ Asegúrate que estos entry IDs SON LOS CORRECTOS
    payload_cf = {
        "entry.1771049143": estilo_lata_cf,
        "entry.251410382": str(cantidad_lata_cf),
        "entry.122481390": lote_lata_cf,
        "entry.268194084": responsable
    }

    response = requests.post(form_cf_url, data=payload_cf)

    if response.status_code in [200, 302]:
        st.success("✅ Ingreso de latas al cuarto frío guardado correctamente")
        st.balloons()
    else:
        st.error(f"❌ Error al guardar ingreso al cuarto frío. Código: {response.status_code}")

# BUSCAR REGISTROS
# ----------------------------------------
st.markdown("---")
st.markdown("<h2 style='color:#fff3aa;'>🔍 Buscar Barriles</h2>", unsafe_allow_html=True)

try:
    df_search = pd.read_csv(url_datos)
    df_search.columns = df_search.columns.str.strip()

    filtro_codigo = st.text_input("🔎 Buscar por código de barril")
    filtro_cliente = st.text_input("🔎 Buscar por cliente")
    filtro_estado = st.selectbox("🔎 Buscar por estado", ["", "Despachado", "Lavado en bodega", "Sucio", "En cuarto frío"])

    df_filtrado = df_search.copy()
    if filtro_codigo:
        df_filtrado = df_filtrado[df_filtrado["Código"].astype(str).str.contains(filtro_codigo)]
    if filtro_cliente:
        df_filtrado = df_filtrado[df_filtrado["Cliente"].astype(str).str.contains(filtro_cliente, case=False)]
    if filtro_estado:
        df_filtrado = df_filtrado[df_filtrado["Estado"] == filtro_estado]

    if not df_filtrado.empty:
        st.dataframe(df_filtrado[["Marca temporal","Código", "Estilo", "Estado", "Cliente", "Responsable", "Observaciones"]])
    else:
        st.warning("⚠️ No se encontraron registros con los filtros aplicados.")

except Exception as e:
    st.error(f"⚠️ No se pudo cargar la hoja de búsqueda: {e}")
           
# ------------------ FORMULARIO GENERAL ------------------
tipo_devolucion = st.selectbox("Selecciona tipo de devolución:", ["", "Barril", "Latas"])
cliente = st.text_input("Cliente que devuelve")
responsable = st.text_input("Responsable que recibe")
observaciones = st.text_area("Observaciones (opcional)")
fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

# ------------------ CAMPOS ESPECÍFICOS ------------------
if tipo_devolucion == "Barril":
    codigo_barril = st.text_input("Código del barril")
    estilo_cerveza = st.text_input("Estilo de cerveza (opcional)")
    lote_producto = st.text_input("Lote del producto (opcional)")

elif tipo_devolucion == "Latas":
    cantidad_latas = st.number_input("Cantidad de latas devueltas", min_value=1, step=1)
    estilo_cerveza = st.text_input("Estilo de cerveza")
    lote_latas = st.text_input("Lote de producto")

# ------------------ BOTÓN DE ENVÍO ------------------
if st.button("Registrar devolución"):
    if tipo_devolucion == "Barril":
        try:
            # ✅ URL del formulario de Google Forms para barriles
            url_form_barril = "https://docs.google.com/forms/d/e/1FAIpQLSedFQmZuDdVY_cqU9WdiWCTBWCCh1NosPnD891QifQKqaeUfA/formResponse"

            # 📤 Datos a enviar (asegúrate que los entry.xxx coincidan con tu formulario real)
            form_data_barril = {
                "entry.311770370": codigo_barril,          # Código del barril
                "entry.1283669263": estilo_cerveza,        # Estilo (opcional)
                "entry.1545499818": "Devolución",          # Estado fijo al devolver
                "entry.91059345": cliente,                 # Cliente
                "entry.1661747572": responsable,           # Responsable
                "entry.1465957833": observaciones,         # Observaciones
                "entry.1234567890": lote_producto,         # Lote (opcional, cambia si tienes otro ID)
                "entry.1437332932": lote_producto          # Lote repetido si tu formulario lo necesita
            }

            response = requests.post(url_form_barril, data=form_data_barril)

            if response.status_code in [200, 302]:
                st.success("✅ Devolución de barril registrada correctamente.")
            else:
                st.warning(f"⚠️ Error al enviar devolución. Código: {response.status_code}")
        except Exception as e:
            st.error(f"❌ Error al registrar devolución de barril: {e}")

    elif tipo_devolucion == "Latas":
        try:
            url_form_latas = "https://docs.google.com/forms/d/e/1FAIpQLSerxxOI1npXAptsa3nvNNBFHYBLV9OMMX-4-Xlhz-VOmitRfQ/formResponse"

            form_data_latas = {
                "entry.457965266": str(cantidad_latas),        # Cantidad
                "entry.689047838": estilo_cerveza,             # Estilo
                "entry.2096096606": lote_latas,                # Lote
                "entry.1478892985": cliente,                   # Cliente
                "entry.1774006398": responsable,               # Responsable
                "entry.1179145668": "Devolución"               # NUEVO CAMPO - Tipo de movimiento
            }

            response = requests.post(url_form_latas, data=form_data_latas)
            if response.status_code in [200, 302]:
                st.success("✅ Devolución de latas registrada correctamente.")
                st.toast("🍺 Devolución registrada")
            else:
                st.warning(f"⚠️ Error al enviar devolución. Código: {response.status_code}")
        except Exception as e:
            st.error(f"❌ Error al registrar devolución de latas: {e}")

# ------------------ REGISTRO DE BAJA DE PRODUCTO ------------------
st.markdown("---")
st.title("🗑️ Registro de Baja de Producto (Barril o Latas)")

tipo_baja = st.selectbox("Selecciona tipo de baja:", ["", "Barril", "Latas"])
responsable_baja = st.text_input("Responsable que da de baja")
observaciones_baja = st.text_area("Observaciones de baja (opcional)")
fecha_actual_baja = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

# Campos específicos
if tipo_baja == "Barril":
    codigo_barril_baja = st.text_input("Código del barril a dar de baja")
    estilo_baja_barril = st.text_input("Estilo de cerveza (opcional)")
    lote_barril_baja = st.text_input("Lote del producto (opcional)")

elif tipo_baja == "Latas":
    cantidad_latas_baja = st.number_input("Cantidad de latas a dar de baja", min_value=1, step=1)
    estilo_latas_baja = st.text_input("Estilo de cerveza")
    lote_latas_baja = st.text_input("Lote del producto")

# Botón para registrar baja
if st.button("Registrar Baja de Producto"):

    if tipo_baja == "Barril":
        try:
            url_form_barril_baja = "https://docs.google.com/forms/d/e/1FAIpQLSedFQmZuDdVY_cqU9WdiWCTBWCCh1NosPnD891QifQKqaeUfA/formResponse"

            form_data_baja_barril = {
                "entry.311770370": codigo_barril_baja,                      # Código del barril
                "entry.1283669263": estilo_baja_barril,                    # Estilo
                "entry.1545499818": "Baja",                                # Estado automático "Baja"
                "entry.1661747572": responsable_baja,                      # Responsable
                "entry.1465957833": observaciones_baja,                    # Observaciones
                "entry.1234567890": lote_barril_baja,                      # Lote (opcional)
                "entry.1437332932": lote_barril_baja                       # Repetido si es necesario
            }

            response = requests.post(url_form_barril_baja, data=form_data_baja_barril)
            if response.status_code in [200, 302]:
                st.success("✅ Baja de barril registrada correctamente.")
                st.balloons()
            else:
                st.warning(f"⚠️ Error al registrar baja. Código: {response.status_code}")
        except Exception as e:
            st.error(f"❌ Error al registrar baja de barril: {e}")

    elif tipo_baja == "Latas":
        try:
            url_form_latas_baja = "https://docs.google.com/forms/d/e/1FAIpQLSerxxOI1npXAptsa3nvNNBFHYBLV9OMMX-4-Xlhz-VOmitRfQ/formResponse"

            form_data_baja_latas = {
                "entry.457965266": str(cantidad_latas_baja),              # Cantidad
                "entry.689047838": estilo_latas_baja,                     # Estilo
                "entry.2096096606": lote_latas_baja,                      # Lote
                "entry.1774006398": responsable_baja,                     # Responsable
                "entry.1179145668": "Baja"                                # Campo adicional para indicar baja
            }

            response = requests.post(url_form_latas_baja, data=form_data_baja_latas)
            if response.status_code in [200, 302]:
                st.success("✅ Baja de latas registrada correctamente.")
                st.snow()
            else:
                st.warning(f"⚠️ Error al registrar baja. Código: {response.status_code}")
        except Exception as e:
            st.error(f"❌ Error al registrar baja de latas: {e}")
